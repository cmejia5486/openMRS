#!/usr/bin/env python3

from __future__ import annotations

import unicodedata
import json
import os
import re
import sys
import time
import hashlib
import platform
from datetime import datetime, timezone
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import openpyxl

try:
    from lib.ai_runtime import AIRuntime
except Exception:
    AIRuntime = None  # type: ignore

try:
    from pydantic import BaseModel
except Exception:
    BaseModel = None  # type: ignore

try:
    from lib.prompt_telemetry import record_prompt_call
    from lib.prompt_registry import get_prompt_contract
except Exception:
    def record_prompt_call(**kwargs: Any) -> str:  # type: ignore
        return str(kwargs.get("prompt_call_id") or "")
    def get_prompt_contract(prompt_id: str, *, required: bool = True) -> Dict[str, Any]:  # type: ignore
        if required:
            raise KeyError(prompt_id)
        return {}


def _first_nonempty_env(*names: str) -> str:
    for name in names:
        value = os.getenv(name, "").strip()
        if value:
            return value
    return ""


def _repo_root() -> Path:
    # scripts/ai_security_audit_requirements_excel.py -> repository root
    try:
        return Path(__file__).resolve().parents[1]
    except Exception:
        return Path.cwd()


def _resolve_data_dir() -> Path:
    """Resolve a portable working data directory.

    Priority:
    1. Explicit project variables: VISION360_DATA_DIR, AUDIT_DATA_DIR, SECURITY_AUDIT_DATA_DIR.
    2. GitHub Actions runner temp directory: RUNNER_TEMP/vision360-data.
    3. Repository-local fallback: <repo>/.vision360-data.

    This avoids binding the script to any operating-system-specific absolute path.
    """
    explicit = _first_nonempty_env("VISION360_DATA_DIR", "AUDIT_DATA_DIR", "SECURITY_AUDIT_DATA_DIR")
    if explicit:
        return Path(explicit).expanduser()

    runner_temp = os.getenv("RUNNER_TEMP", "").strip()
    if runner_temp:
        return Path(runner_temp).expanduser() / "vision360-data"

    return _repo_root() / ".vision360-data"


def _resolve_path(env_name: str, default_filename: str) -> Path:
    explicit = os.getenv(env_name, "").strip()
    if explicit:
        return Path(explicit).expanduser()
    return DATA_DIR / default_filename


DATA_DIR = _resolve_data_dir()
FINGERPRINT_PATH = _resolve_path("VISION360_FINGERPRINT_PATH", "vision360_fingerprint.json")
REQUISITES_PATH = _resolve_path("REQUISITES_PATH", "requisites.json")
OUTPUT_XLSX_PATH = _resolve_path("SECURITY_AUDIT_XLSX_PATH", "security_audit_requirements.xlsx")
RUN_METRICS_XLSX_PATH = _resolve_path("RUN_METRICS_XLSX_PATH", "run-metrics.xlsx")

RUN_METRICS: Dict[str, List[Dict[str, Any]]] = {
    "llm_calls": [],
    "llm_items": [],
    "adjudication_by_requirement": [],
    "adjudication_by_flag": [],
}

NEGATIVE_RISK_TOKENS = [
    "insecure", "unsafe", "weak", "debug", "debuggable", "cleartext", "allow_clear_text",
    "trust_all", "accept_all", "ignore_ssl", "skip_verification", "bypass", "hardcoded",
    "leak", "plaintext", "world_readable", "world_writable", "sha1", "md5",
    "debug_certificate", "janus", "v1_signature", "exported_true", "backup_enabled_true",
    "http_based",
]

APPLICABILITY_TOKENS = [
    "components", "present", "detected", "uses_", "is_used", "feature", "webview_components",
]

PASSWORD_HASHING_POSITIVE_IDS = {"has_password_hashing_uses_salts", "has_password_hashing_uses_kdf"}

MALWARE_REQ_TOKENS = [
    "malware", "adware", "virus", "trojan", "spyware", "ransomware", "malicious code", "malicious",
]
MALWARE_FLAG_TOKENS = ["malware", "adware", "virus", "trojan", "spyware", "ransomware", "malicious"]

OVERRIDE_SCOPE_FLAG_IDS = {
    "has_org_notifies_users_of_security_updates",
    "has_manifest_allow_clear_text_traffic_true",
    "has_uses_os_level_update_mechanisms",
    "has_android_dynamic_code_loading",
    "has_webview_remote_content",
    "has_soap_uses_mutual_tls",
    "has_defined_certificate_management_policy",
    "has_defined_identity_lifecycle_policy",
    "has_webview_components",
    "has_webview_javascript",
    "has_webview_file_scheme",
    "has_insecure_http_based_webview_communication",
    "has_webview_javascript_interface_limited_to_trusted_content",
    "has_soap_api_usage",
    "has_proper_ws_security_headers",
    "has_soap_message_level_encryption",
    "has_soap_message_level_signatures",
    "has_soap_prevents_replay_attacks",
    "has_saml_based_sso",
    "has_soap_validates_saml_token_expiry",
    "has_uses_xml_signatures",
    "has_uses_xml_encryption",
    "has_soap_uses_strict_schema_validation",
    "has_content_provider_actively_exposed",
    "has_manifest_custom_permission_defined",
}

GATE_FLAG_IDS = {
    "has_webview_components",
    "has_webview_remote_content",
    "has_android_dynamic_code_loading",
    "has_soap_api_usage",
    "has_saml_based_sso",
    "has_content_provider_actively_exposed",
    "has_manifest_custom_permission_defined",
    "has_uses_os_level_update_mechanisms",
    "has_org_notifies_users_of_security_updates",
    "has_defined_certificate_management_policy",
    "has_defined_identity_lifecycle_policy",
}


def _json_decode_error_details(e: json.JSONDecodeError) -> str:
    return f"{type(e).__name__}: {e.msg} (line {e.lineno}, col {e.colno})"


def load_json_with_one_repair(path: Path) -> Any:
    raw = path.read_text(encoding="utf-8", errors="replace")
    try:
        return json.loads(raw)
    except json.JSONDecodeError as e:
        print(f"Error parsing {path.name}: {_json_decode_error_details(e)}", file=sys.stderr)
        patched: Optional[str] = None
        if "Extra data" in e.msg:
            patched = raw[: e.pos].rstrip()
        elif raw.lstrip().startswith("{{"):
            s = raw.lstrip()
            idx = raw.find(s)
            patched = raw[:idx] + s[1:]
        if patched is None:
            raise
        return json.loads(patched)


def normalize_requirements(data: Any) -> List[Dict[str, Any]]:
    if isinstance(data, list):
        return data
    if isinstance(data, dict) and isinstance(data.get("requirements"), list):
        return data["requirements"]
    raise ValueError("requisites.json must be a JSON array of requirements (or an object with a 'requirements' array).")


def normalize_fingerprint_flags(data: Any) -> List[Dict[str, Any]]:
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        if isinstance(data.get("flags"), list):
            return data["flags"]
        for k in ("results", "items"):
            if isinstance(data.get(k), list):
                return data[k]
    raise ValueError("vision360_fingerprint.json does not contain a recognizable list of flags.")


def is_prohibitive(req_desc: str) -> bool:
    s = req_desc.lower()
    return ("must not" in s) or ("shall not" in s) or ("should not" in s)


def is_conditional(req_desc: str) -> bool:
    s = req_desc.lower()
    return any(re.search(p, s) for p in [r"\bif\b", r"\bwhen\b", r"where applicable", r"on older versions"])


def req_mentions_malware(req_desc: str) -> bool:
    s = req_desc.lower()
    return any(t in s for t in MALWARE_REQ_TOKENS)


def classify_flag_for_requirement(flag_id: str, flag_title: str, req_desc: str) -> str:
    fid = (flag_id or "").lower()
    title = (flag_title or "").lower()
    if flag_id in PASSWORD_HASHING_POSITIVE_IDS:
        return "POSITIVE_CONTROL"
    if req_mentions_malware(req_desc) and any(t in fid or t in title for t in MALWARE_FLAG_TOKENS):
        return "NEGATIVE_RISK"
    if any(tok in fid or tok in title for tok in NEGATIVE_RISK_TOKENS):
        return "NEGATIVE_RISK"
    if any(tok in fid or tok in title for tok in APPLICABILITY_TOKENS):
        return "APPLICABILITY"
    return "POSITIVE_CONTROL"


def parse_summary_normalized(summary: Any) -> str:
    if not isinstance(summary, str) or not summary.strip():
        return "NA"
    s = summary.strip()
    v = s.split("=")[-1].strip() if "=" in s else s
    v_up = v.upper().replace(" ", "")
    if v_up in ("YES", "Y"):
        return "YES"
    if v_up in ("NO", "N"):
        return "NO"
    if v_up in ("NA", "N/A", "UNKNOWN", "NOT_APPLICABLE", "NOTAPPLICABLE"):
        return "NA"
    m = re.search(r"\b(YES|NO|NA|N/A|UNKNOWN|NOT_APPLICABLE)\b", s.upper())
    if m:
        tok = m.group(1)
        return "NA" if tok in ("NA", "N/A", "UNKNOWN", "NOT_APPLICABLE") else tok
    return "NA"


def eval_against_expected(observed: str, expected: str) -> str:
    if observed == "NA":
        return "UNKNOWN"
    if expected == "YES":
        return "SUPPORT" if observed == "YES" else "CONTRADICT"
    if expected == "NO":
        return "SUPPORT" if observed == "NO" else "CONTRADICT"
    return "UNKNOWN"


@dataclass
class FlagEvidence:
    id: str
    title: str
    state: str
    summary: str
    summary_norm: str
    notes: str
    evidence_count: int
    classification: str
    expected: Optional[str]
    outcome: str


@dataclass
class RequirementAudit:
    puid: str
    description_en: str
    result: str
    flags_used: List[str]
    justification_en: str


def build_flag_evidence(flag_obj: Optional[Dict[str, Any]], flag_id: str, classification: str, expected: Optional[str]) -> FlagEvidence:
    if not flag_obj:
        return FlagEvidence(flag_id, "", "missing", "", "NA", "flag not present in fingerprint", 0, classification, expected, "MISSING")
    app_verdict = flag_obj.get("app_verdict") or {}
    summary = app_verdict.get("summary", "")
    summary_norm = parse_summary_normalized(summary)
    state = str(app_verdict.get("state", "") or "")
    notes = str(app_verdict.get("notes", "") or "")
    evidence_count = int(app_verdict.get("evidence_count", 0) or 0)
    if classification == "APPLICABILITY":
        outcome = "NA_OR_GATE"
    else:
        outcome = eval_against_expected(summary_norm, expected or "YES")
    return FlagEvidence(str(flag_obj.get("id") or flag_id), str(flag_obj.get("title") or ""), state, str(summary), summary_norm, notes, evidence_count, classification, expected, outcome)


def compute_override_scenario_activated(flag_evidences: List[FlagEvidence], flag_ids: List[str]) -> Tuple[bool, List[FlagEvidence]]:
    gate_evs = [fe for fe in flag_evidences if fe.id in flag_ids and fe.id in GATE_FLAG_IDS]
    return any(fe.summary_norm == "YES" for fe in gate_evs), gate_evs


def compute_conditional_scenario_activated(flag_evidences: List[FlagEvidence]) -> Optional[bool]:
    app_flags = [fe for fe in flag_evidences if fe.classification == "APPLICABILITY"]
    if not app_flags:
        return None
    return any(fe.summary_norm == "YES" for fe in app_flags)


def _split_flags_string(flags_str: str) -> List[str]:
    s = (flags_str or "").strip()
    if not s:
        return []
    if s.startswith("[") and s.endswith("]"):
        try:
            obj = json.loads(s)
            if isinstance(obj, list):
                return [str(x).strip() for x in obj if str(x).strip()]
        except Exception:
            pass
    return [p.strip() for p in re.split(r"[,\n;]+", s) if p.strip()]


def extract_req_fields(req_obj: Dict[str, Any]) -> Tuple[str, str, List[str]]:
    puid = str(req_obj.get("PUID") or req_obj.get("id") or "").strip()
    desc = str(req_obj.get("Requirement description") or req_obj.get("Description") or req_obj.get("Descripcion") or "").strip()
    flags = req_obj.get("Flags", [])
    if isinstance(flags, str):
        flags_list = _split_flags_string(flags)
    elif isinstance(flags, list):
        flags_list = [str(x).strip() for x in flags if str(x).strip()]
    else:
        flags_list = []
    return puid, desc, flags_list


def audit_requirement(puid: str, desc: str, flag_ids: List[str], flags_by_id: Dict[str, Dict[str, Any]]) -> Tuple[str, List[FlagEvidence], Dict[str, Any]]:
    prohibitive = is_prohibitive(desc)
    conditional = is_conditional(desc)
    flag_evs: List[FlagEvidence] = []
    for fid in flag_ids:
        fobj = flags_by_id.get(fid)
        title = str((fobj or {}).get("title") or "")
        classification = classify_flag_for_requirement(fid, title, desc)
        expected: Optional[str] = None
        if classification == "POSITIVE_CONTROL":
            expected = "YES"
        elif classification == "NEGATIVE_RISK":
            expected = "NO"
        flag_evs.append(build_flag_evidence(fobj, fid, classification, expected))

    override_used = bool(flag_ids) and set(flag_ids).issubset(OVERRIDE_SCOPE_FLAG_IDS)
    override_scenario_activated: Optional[bool] = None
    gate_flags: List[FlagEvidence] = []
    conditional_scenario_activated: Optional[bool] = None

    has_negative_risk_yes = any((fe.classification == "NEGATIVE_RISK" and fe.summary_norm == "YES") for fe in flag_evs if fe.outcome != "MISSING")

    if override_used:
        override_scenario_activated, gate_flags = compute_override_scenario_activated(flag_evs, flag_ids)
        if override_scenario_activated is False:
            result = ("no" if has_negative_risk_yes else "yes") if prohibitive else ("no" if has_negative_risk_yes else "n/a")
            meta = dict(prohibitive=prohibitive, conditional=conditional, override_used=True, override_scenario_activated=override_scenario_activated, gate_flags=[ge.id for ge in gate_flags], conditional_scenario_activated=None)
            return result, flag_evs, meta

    if conditional:
        conditional_scenario_activated = compute_conditional_scenario_activated(flag_evs)

    non_app = [fe for fe in flag_evs if fe.classification != "APPLICABILITY"]
    any_contradict = any(fe.outcome == "CONTRADICT" for fe in non_app)
    all_support = (len(non_app) > 0) and all(fe.outcome == "SUPPORT" for fe in non_app)
    any_unknown = any(fe.outcome in ("UNKNOWN", "MISSING") for fe in non_app)

    if any_contradict:
        result = "no"
    elif all_support and not any_unknown:
        result = "yes"
    else:
        if conditional and (conditional_scenario_activated is False):
            result = "yes" if prohibitive else "n/a"
        else:
            result = "n/a"

    meta = dict(prohibitive=prohibitive, conditional=conditional, override_used=override_used, override_scenario_activated=override_scenario_activated, gate_flags=[ge.id for ge in gate_flags], conditional_scenario_activated=conditional_scenario_activated)
    return result, flag_evs, meta


def _bool_metric(value: bool) -> str:
    return "true" if bool(value) else "false"


def _join_ids(items: List[FlagEvidence]) -> str:
    return ", ".join(fe.id for fe in items if fe and fe.id)


def _adjudication_decision_basis(result: str, flag_evs: List[FlagEvidence], meta: Dict[str, Any]) -> str:
    non_app = [fe for fe in flag_evs if fe.classification != "APPLICABILITY"]
    any_contradict = any(fe.outcome == "CONTRADICT" for fe in non_app)
    any_unknown = any(fe.outcome in ("UNKNOWN", "MISSING") for fe in non_app)
    all_support = bool(non_app) and all(fe.outcome == "SUPPORT" for fe in non_app)
    if any_contradict:
        return "risk_precedence"
    if meta.get("override_used") and meta.get("override_scenario_activated") is False:
        return "override_scope_not_activated"
    if meta.get("conditional") and meta.get("conditional_scenario_activated") is False:
        return "conditional_not_activated"
    if result == "yes" and all_support and not any_unknown:
        return "full_support"
    if any_unknown or any(fe.classification == "APPLICABILITY" for fe in flag_evs):
        return "partial_or_unknown"
    if not non_app:
        return "no_mapped_non_applicability_flags"
    return "default_rule"


def _record_adjudication_diagnostics(
    *,
    puid: str,
    result: str,
    flag_ids: List[str],
    flag_evs: List[FlagEvidence],
    meta: Dict[str, Any],
) -> None:
    """Record requirement-level and flag-level adjudication diagnostics.

    These diagnostics are transparency metrics for auditability. They do not
    alter the deterministic yes/no/n/a result. A contradiction here means a
    mapped non-applicability flag observed a value incompatible with the secure
    posture expected for that requirement. The risk-precedence rule is already
    the existing audit behavior: any contradictory mapped signal prevents a
    compliant result.
    """
    non_app = [fe for fe in flag_evs if fe.classification != "APPLICABILITY"]
    support = [fe for fe in non_app if fe.outcome == "SUPPORT"]
    contradict = [fe for fe in non_app if fe.outcome == "CONTRADICT"]
    unknown = [fe for fe in non_app if fe.outcome == "UNKNOWN"]
    missing = [fe for fe in non_app if fe.outcome == "MISSING"]
    applicability = [fe for fe in flag_evs if fe.classification == "APPLICABILITY"]
    evidence_total = sum(int(fe.evidence_count or 0) for fe in flag_evs)
    decision_basis = _adjudication_decision_basis(result, flag_evs, meta)
    contradictory_requirement = bool(contradict)
    risk_precedence_applied = bool(result == "no" and contradictory_requirement)
    partial_evidence = bool(not contradictory_requirement and (unknown or missing) and result in {"n/a", "no"})

    RUN_METRICS["adjudication_by_requirement"].append({
        "puid": puid,
        "result": result,
        "decision_basis": decision_basis,
        "total_mapped_flags": len(flag_ids),
        "support_flag_count": len(support),
        "contradict_flag_count": len(contradict),
        "unknown_flag_count": len(unknown),
        "missing_flag_count": len(missing),
        "applicability_flag_count": len(applicability),
        "evidence_count_total": evidence_total,
        "contradictory_requirement": _bool_metric(contradictory_requirement),
        "risk_precedence_applied": _bool_metric(risk_precedence_applied),
        "partial_evidence": _bool_metric(partial_evidence),
        "conditional": _bool_metric(bool(meta.get("conditional"))),
        "conditional_scenario_activated": "" if meta.get("conditional_scenario_activated") is None else _bool_metric(bool(meta.get("conditional_scenario_activated"))),
        "override_used": _bool_metric(bool(meta.get("override_used"))),
        "override_scenario_activated": "" if meta.get("override_scenario_activated") is None else _bool_metric(bool(meta.get("override_scenario_activated"))),
        "support_flag_ids": _join_ids(support),
        "contradict_flag_ids": _join_ids(contradict),
        "unknown_flag_ids": _join_ids(unknown),
        "missing_flag_ids": _join_ids(missing),
        "applicability_flag_ids": _join_ids(applicability),
    })

    for fe in flag_evs:
        if fe.classification == "APPLICABILITY":
            role = "applicability_gate"
        elif fe.outcome == "SUPPORT":
            role = "supports_expected_posture"
        elif fe.outcome == "CONTRADICT":
            role = "contradicts_expected_posture"
        elif fe.outcome == "MISSING":
            role = "missing_from_fingerprint"
        else:
            role = "unknown_or_non_conclusive"
        RUN_METRICS["adjudication_by_flag"].append({
            "puid": puid,
            "requirement_result": result,
            "flag_id": fe.id,
            "flag_title": fe.title,
            "classification": fe.classification,
            "expected_summary_norm": fe.expected or "",
            "observed_summary_norm": fe.summary_norm,
            "outcome": fe.outcome,
            "decision_role": role,
            "state": fe.state,
            "evidence_count": fe.evidence_count,
            "notes_excerpt": (fe.notes or "")[:500],
        })


def _adjudication_summary_rows() -> List[Tuple[str, Any]]:
    rows = RUN_METRICS.get("adjudication_by_requirement", [])
    flag_rows = RUN_METRICS.get("adjudication_by_flag", [])
    total_requirements = len(rows)
    contradictory_requirement_count = sum(1 for r in rows if str(r.get("contradictory_requirement", "")).lower() == "true")
    risk_precedence_count = sum(1 for r in rows if str(r.get("risk_precedence_applied", "")).lower() == "true")
    partial_evidence_requirement_count = sum(1 for r in rows if str(r.get("partial_evidence", "")).lower() == "true")
    contradictory_flag_reference_count = sum(1 for r in flag_rows if str(r.get("outcome", "")) == "CONTRADICT")
    unknown_flag_reference_count = sum(1 for r in flag_rows if str(r.get("outcome", "")) == "UNKNOWN")
    missing_flag_reference_count = sum(1 for r in flag_rows if str(r.get("outcome", "")) == "MISSING")
    support_flag_reference_count = sum(1 for r in flag_rows if str(r.get("outcome", "")) == "SUPPORT")
    applicability_flag_reference_count = sum(1 for r in flag_rows if str(r.get("classification", "")) == "APPLICABILITY")
    return [
        ("adjudicated_requirement_count", total_requirements),
        ("adjudicated_flag_reference_count", len(flag_rows)),
        ("support_flag_reference_count", support_flag_reference_count),
        ("contradictory_requirement_count", contradictory_requirement_count),
        ("contradictory_flag_reference_count", contradictory_flag_reference_count),
        ("risk_precedence_count", risk_precedence_count),
        ("partial_evidence_requirement_count", partial_evidence_requirement_count),
        ("unknown_flag_reference_count", unknown_flag_reference_count),
        ("missing_flag_reference_count", missing_flag_reference_count),
        ("applicability_flag_reference_count", applicability_flag_reference_count),
    ]


class _ResponsesCompat:
    def __init__(self, runtime: Any) -> None:
        self._runtime = runtime

    def create(self, model: Optional[str] = None, input: Any = None, max_output_tokens: Optional[int] = None, reasoning: Optional[dict] = None):
        return self._runtime.create(input=input, model=model, max_output_tokens=max_output_tokens, reasoning=reasoning)

    def parse(self, model: Optional[str] = None, input: Any = None, text_format: Any = None, max_output_tokens: Optional[int] = None, reasoning: Optional[dict] = None):
        return self._runtime.parse(input=input, text_format=text_format, model=model, max_output_tokens=max_output_tokens, reasoning=reasoning)


class _ClientCompat:
    def __init__(self, runtime: Any) -> None:
        self.responses = _ResponsesCompat(runtime)


def openai_client() -> Optional[Any]:
    api_key = os.getenv("OPENAI_API_KEY", "").strip()
    if not api_key or AIRuntime is None:
        return None
    task = os.getenv("AI_TASK", "").strip() or "ai_requirements_excel"
    return _ClientCompat(AIRuntime(task=task, api_key=api_key))


def env_int(name: str, default: int) -> int:
    v = os.getenv(name, "").strip()
    if not v:
        return default
    try:
        return int(v)
    except ValueError:
        return default


def env_bool(name: str, default: bool) -> bool:
    v = os.getenv(name, "").strip().lower()
    if not v:
        return default
    if v in ("1", "true", "yes", "y", "on"):
        return True
    if v in ("0", "false", "no", "n", "off"):
        return False
    return default


def is_local_openai_compatible_endpoint() -> bool:
    api_base = (
        os.getenv("AI_API_BASE", "")
        or os.getenv("OPENAI_API_BASE", "")
        or os.getenv("AI_COMPAT_BASE_URL", "")
    ).strip().lower()
    return bool(api_base) and any(
        host in api_base
        for host in ("localhost", "127.0.0.1", "host.docker.internal")
    )


def should_use_structured_parse(client: Any) -> bool:
    # LM Studio and other local OpenAI-compatible servers often expose enough
    # surface for the SDK call to exist, but they do not reliably return the
    # same structured-output envelope expected by client.responses.parse.
    # For local endpoints, use plain text generation and parse JSON ourselves.
    return hasattr(client.responses, "parse") and not is_local_openai_compatible_endpoint()


def extract_json_object_from_model_output(text: str) -> Dict[str, Any]:
    text = (text or "").strip()
    if not text:
        raise ValueError("Model returned empty text")

    # Direct JSON.
    try:
        obj = json.loads(text)
        if isinstance(obj, dict):
            return obj
    except Exception:
        pass

    # JSON inside Markdown fence.
    fenced = re.search(
        r"```(?:json)?\s*(\{.*?\})\s*```",
        text,
        flags=re.DOTALL | re.IGNORECASE,
    )
    if fenced:
        obj = json.loads(fenced.group(1))
        if isinstance(obj, dict):
            return obj

    # Remove common fence wrappers if the model used an unterminated fence.
    cleaned = re.sub(r"^```(?:json)?\s*", "", text, flags=re.IGNORECASE).strip()
    cleaned = re.sub(r"\s*```$", "", cleaned).strip()
    try:
        obj = json.loads(cleaned)
        if isinstance(obj, dict):
            return obj
    except Exception:
        pass

    # Find the first balanced JSON object anywhere in the output.
    start = text.find("{")
    if start < 0:
        preview = text[:500].replace("\n", "\\n")
        raise ValueError(f"No JSON object found in model output. Preview: {preview}")

    depth = 0
    in_string = False
    escape = False

    for i in range(start, len(text)):
        ch = text[i]

        if escape:
            escape = False
            continue

        if ch == "\\":
            escape = True
            continue

        if ch == '"':
            in_string = not in_string
            continue

        if in_string:
            continue

        if ch == "{":
            depth += 1
        elif ch == "}":
            depth -= 1
            if depth == 0:
                candidate = text[start:i + 1]
                obj = json.loads(candidate)
                if isinstance(obj, dict):
                    return obj

    preview = text[:500].replace("\n", "\\n")
    raise ValueError(f"No complete JSON object found in model output. Preview: {preview}")


def _normalize_typography(s: str) -> str:
    repl = {"\u2018": "'", "\u2019": "'", "\u201C": '"', "\u201D": '"', "\u2013": "-", "\u2014": "-", "\u00A0": " "}
    return "".join(repl.get(ch, ch) for ch in s)


def looks_non_english(text: str) -> bool:
    if not text or not text.strip():
        return False
    s = _normalize_typography(text.strip())
    for ch in s:
        if ord(ch) <= 127:
            continue
        cat = unicodedata.category(ch)
        if cat.startswith("L") or cat.startswith("M"):
            return True
    spanish_markers = [" el ", " la ", " los ", " las ", " de ", " del ", " para ", " y ", " o ", " debe ", " cuando ", " donde ", " aplicacion ", " seguridad ", " requisito ", " descripcion ", " evidencias "]
    s_low = f" {s.lower()} "
    return any(m in s_low for m in spanish_markers)



def _sha256_text(value: str) -> str:
    return hashlib.sha256(str(value or "").encode("utf-8", errors="replace")).hexdigest()


def _sha256_file(path: Path) -> str:
    try:
        if not path or not path.is_file():
            return ""
        h = hashlib.sha256()
        with path.open("rb") as fh:
            for chunk in iter(lambda: fh.read(1024 * 1024), b""):
                h.update(chunk)
        return h.hexdigest()
    except Exception:
        return ""


def _stable_json_hash(value: Any) -> str:
    try:
        payload = json.dumps(value, sort_keys=True, ensure_ascii=False, default=str)
    except Exception:
        payload = str(value)
    return _sha256_text(payload)


def _now_utc_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _env_snapshot(keys: List[str]) -> Dict[str, str]:
    return {key: os.getenv(key, "") for key in keys}


def _llm_config_snapshot(batch_size: int | None = None) -> Dict[str, Any]:
    keys = [
        "AI_TASK", "AI_PROFILE", "AI_PROVIDER", "AI_MODEL", "AI_LITELLM_MODEL",
        "AI_API_BASE", "AI_REASONING_EFFORT", "AI_MAX_OUTPUT_TOKENS", "AI_BATCH_SIZE",
        "OPENAI_MODEL", "OPENAI_REASONING_EFFORT", "OPENAI_MAX_OUTPUT_TOKENS",
        "OPENAI_BATCH_SIZE", "OPENAI_TIMEOUT_S", "USE_OPENAI_JUSTIFICATIONS",
        "STRICT_ENGLISH_OUTPUT",
    ]
    snap: Dict[str, Any] = _env_snapshot(keys)
    if batch_size is not None:
        snap["effective_batch_size"] = batch_size
    return snap


def _record_llm_call(
    *,
    call_type: str,
    expected_items: int,
    received_items: int,
    json_valid: bool,
    schema_valid: bool,
    retry_count: int,
    elapsed_s: float,
    model: str,
    max_tokens: int,
    parse_route: bool,
    error: str = "",
    prompt_call_id: str = "",
    prompt_id: str = "",
    prompt_scope: str = "",
    prompt_category: str = "",
) -> None:
    call_id = f"LLM-{len(RUN_METRICS['llm_calls']) + 1:04d}"
    RUN_METRICS["llm_calls"].append({
        "llm_call_id": call_id,
        "prompt_call_id": prompt_call_id,
        "prompt_id": prompt_id,
        "prompt_scope": prompt_scope,
        "prompt_category": prompt_category,
        "call_type": call_type,
        "expected_items": int(expected_items),
        "received_items": int(received_items),
        "json_valid": bool(json_valid),
        "schema_valid": bool(schema_valid),
        "retry_count": int(max(0, retry_count)),
        "elapsed_s": round(float(elapsed_s), 3),
        "model": model,
        "max_output_tokens": int(max_tokens),
        "parse_route": bool(parse_route),
        "error": str(error or "")[:500],
    })


def _append_llm_item(row: Dict[str, Any]) -> None:
    RUN_METRICS["llm_items"].append(dict(row))


def _row_hash_for_requirement(puid: str, result: str, flags_used: List[str]) -> str:
    return _stable_json_hash({"puid": puid, "result": result, "flags_used": list(flags_used or [])})


def _auto_width(ws: Any, max_width: int = 72) -> None:
    for col in ws.columns:
        max_len = 0
        letter = col[0].column_letter
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max(max_len + 2, 10), max_width)


def _append_kv_sheet(wb: Any, title: str, rows: List[Tuple[str, Any]]) -> Any:
    ws = wb.create_sheet(title)
    ws.append(["key", "value"])
    for key, value in rows:
        ws.append([key, value])
    _auto_width(ws)
    return ws


def _append_dict_rows_sheet(wb: Any, title: str, rows: List[Dict[str, Any]], headers: List[str]) -> Any:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])
    _auto_width(ws)
    return ws


def _style_run_metrics_workbook(wb: Any) -> None:
    """Apply a light professional style to the raw per-run metrics workbook.

    Styling is presentation-only. It does not add formulas, alter result values,
    or change the requirement-level yes/no/n/a adjudication.
    """
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter

    theme_navy = "17365D"
    theme_light_blue = "EEF5FB"
    theme_light = "F7F9FC"
    theme_text = "1F2933"
    header_fill = PatternFill("solid", fgColor=theme_navy)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    body_font = Font(name="Arial", color=theme_text, size=9)
    key_fill = PatternFill("solid", fgColor=theme_light_blue)
    alt_fill = PatternFill("solid", fgColor=theme_light)
    side = Side(style="thin", color="D0D7DE")
    border = Border(left=side, right=side, top=side, bottom=side)
    tab_colors = {
        "run_summary": theme_navy,
        "input_hashes": "5B9BD5",
        "compliance_export": "70AD47",
        "adjudication_summary": "FFC000",
        "adjudication_by_requirement": "FFC000",
        "adjudication_by_flag": "FFC000",
        "llm_calls": "ED7D31",
        "llm_items": "A5A5A5",
    }

    wb.properties.title = "Run Metrics"
    wb.properties.subject = "Per-execution audit telemetry"
    wb.properties.creator = "mSEC-AM workflow"

    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        ws.sheet_properties.tabColor = tab_colors.get(ws.title, theme_navy)
        max_row = ws.max_row
        max_col = ws.max_column
        if max_row < 1 or max_col < 1:
            continue
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 24
        for row_idx in range(2, max_row + 1):
            for col_idx in range(1, max_col + 1):
                cell = ws.cell(row_idx, col_idx)
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
                if ws.title == "run_summary" and col_idx == 1:
                    cell.fill = key_fill
                    cell.font = Font(name="Arial", bold=True, color=theme_text, size=9)
                elif row_idx % 2 == 0:
                    cell.fill = alt_fill
                if isinstance(cell.value, bool):
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = "#,##0.000" if isinstance(cell.value, float) and abs(cell.value - round(cell.value)) > 1e-9 else "#,##0"
        ws.auto_filter.ref = ws.dimensions
        try:
            ref = f"A1:{get_column_letter(max_col)}{max_row}"
            table_name = "tbl_" + "".join(ch if ch.isalnum() else "_" for ch in ws.title.lower())[:230]
            table = Table(displayName=table_name, ref=ref)
            table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
            ws.add_table(table)
        except Exception:
            pass
        for col_idx in range(1, max_col + 1):
            letter = get_column_letter(col_idx)
            header = str(ws.cell(1, col_idx).value or "").lower()
            max_len = len(str(ws.cell(1, col_idx).value or ""))
            for row_idx in range(2, min(max_row, 200) + 1):
                max_len = max(max_len, len(str(ws.cell(row_idx, col_idx).value or "")))
            if "hash" in header or "sha" in header:
                width = 34
            elif "path" in header or "flags" in header or "error" in header:
                width = 42
            elif "justification" in header:
                width = 36
            elif "puid" in header or "id" in header:
                width = 22
            else:
                width = min(max(max_len + 2, 12), 30)
            ws.column_dimensions[letter].width = width


def _write_run_metrics_raw_xlsx(
    *,
    audits: List[RequirementAudit],
    counts: Dict[str, int],
    run_started_at: float,
    batch_size: int,
    total_requirements: int,
    strict_english: bool,
    use_openai_justifications: bool,
    non_english_requirements_detected: int,
) -> None:
    """Write per-execution raw telemetry for downstream repeated-run analysis.

    This function is intentionally non-invasive: it does not change requirement
    adjudication, workbook verdicts, or AI-generated justifications. It only
    exports raw execution telemetry and requirement-level hashes.
    """
    RUN_METRICS_XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)

    llm_snapshot = _llm_config_snapshot(batch_size=batch_size)
    llm_config_hash = _stable_json_hash(llm_snapshot)
    compliance_rows: List[Dict[str, Any]] = []
    for a in audits:
        flags_joined = ", ".join(a.flags_used)
        row_hash = _row_hash_for_requirement(a.puid, a.result, a.flags_used)
        compliance_rows.append({
            "puid": a.puid,
            "result": a.result,
            "flags_used": flags_joined,
            "flags_count": len(a.flags_used),
            "justification_length_chars": len(a.justification_en or ""),
            "row_hash": row_hash,
        })
    compliance_matrix_hash = _stable_json_hash([r["row_hash"] for r in compliance_rows])

    input_hash_rows = [
        {"input_name": "vision360_fingerprint", "path": str(FINGERPRINT_PATH), "sha256": _sha256_file(FINGERPRINT_PATH)},
        {"input_name": "requisites", "path": str(REQUISITES_PATH), "sha256": _sha256_file(REQUISITES_PATH)},
        {"input_name": "security_audit_requirements", "path": str(OUTPUT_XLSX_PATH), "sha256": _sha256_file(OUTPUT_XLSX_PATH)},
    ]

    elapsed_s = round(time.time() - run_started_at, 3)
    run_summary_rows = [
        ("run_id", os.getenv("GITHUB_RUN_ID", "") or f"local-{datetime.now(timezone.utc).strftime('%Y%m%dT%H%M%SZ')}") ,
        ("run_attempt", os.getenv("GITHUB_RUN_ATTEMPT", "")),
        ("workflow", os.getenv("GITHUB_WORKFLOW", "")),
        ("job", os.getenv("GITHUB_JOB", "")),
        ("commit_sha", os.getenv("GITHUB_SHA", "")),
        ("ref", os.getenv("GITHUB_REF", "")),
        ("repository", os.getenv("GITHUB_REPOSITORY", "")),
        ("runner_os", os.getenv("RUNNER_OS", platform.system())),
        ("python_version", platform.python_version()),
        ("generated_at_utc", _now_utc_iso()),
        ("elapsed_s", elapsed_s),
        ("num_requirements", total_requirements),
        ("num_yes", counts.get("yes", 0)),
        ("num_no", counts.get("no", 0)),
        ("num_na", counts.get("n/a", 0)),
    ]
    run_summary_rows.extend(_adjudication_summary_rows())
    run_summary_rows.extend([
        ("batch_size", batch_size),
        ("num_batches", (total_requirements + batch_size - 1) // batch_size if batch_size else 0),
        ("use_openai_justifications", str(bool(use_openai_justifications))),
        ("strict_english_output", str(bool(strict_english))),
        ("requirements_language_gate", "strict_english_catalog" if strict_english else "advisory_english_catalog"),
        ("non_english_requirements_detected", non_english_requirements_detected),
        ("llm_config_hash", llm_config_hash),
        ("compliance_matrix_hash", compliance_matrix_hash),
        ("output_xlsx_path", str(OUTPUT_XLSX_PATH)),
        ("run_metrics_xlsx_path", str(RUN_METRICS_XLSX_PATH)),
    ])
    for key, value in llm_snapshot.items():
        run_summary_rows.append((f"config.{key}", value))

    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)
    _append_kv_sheet(wb, "run_summary", run_summary_rows)
    _append_dict_rows_sheet(wb, "input_hashes", input_hash_rows, ["input_name", "path", "sha256"])
    _append_dict_rows_sheet(wb, "compliance_export", compliance_rows, ["puid", "result", "flags_used", "flags_count", "justification_length_chars", "row_hash"])
    _append_kv_sheet(wb, "adjudication_summary", _adjudication_summary_rows())
    _append_dict_rows_sheet(wb, "adjudication_by_requirement", RUN_METRICS["adjudication_by_requirement"], [
        "puid", "result", "decision_basis", "total_mapped_flags", "support_flag_count",
        "contradict_flag_count", "unknown_flag_count", "missing_flag_count",
        "applicability_flag_count", "evidence_count_total", "contradictory_requirement",
        "risk_precedence_applied", "partial_evidence", "conditional",
        "conditional_scenario_activated", "override_used", "override_scenario_activated",
        "support_flag_ids", "contradict_flag_ids", "unknown_flag_ids", "missing_flag_ids",
        "applicability_flag_ids",
    ])
    _append_dict_rows_sheet(wb, "adjudication_by_flag", RUN_METRICS["adjudication_by_flag"], [
        "puid", "requirement_result", "flag_id", "flag_title", "classification",
        "expected_summary_norm", "observed_summary_norm", "outcome", "decision_role",
        "state", "evidence_count", "notes_excerpt",
    ])
    _append_dict_rows_sheet(wb, "llm_calls", RUN_METRICS["llm_calls"], [
        "llm_call_id", "prompt_call_id", "prompt_id", "prompt_scope", "prompt_category", "call_type",
        "expected_items", "received_items", "json_valid", "schema_valid", "retry_count", "elapsed_s",
        "model", "max_output_tokens", "parse_route", "error",
    ])
    _append_dict_rows_sheet(wb, "llm_items", RUN_METRICS["llm_items"], [
        "run_batch", "puid", "expected_by_llm", "received_from_llm", "field_complete",
        "traceability_ok", "fallback_used", "justification_source", "result", "flags_count",
    ])
    _style_run_metrics_workbook(wb)
    wb.save(RUN_METRICS_XLSX_PATH)
    print(f"[OK] Run metrics workbook generated: {RUN_METRICS_XLSX_PATH}", flush=True)



def _coerce_justification_map(obj: Any, expected_ids: List[str]) -> Dict[str, str]:
    expected_set = set(expected_ids)
    out: Dict[str, str] = {}

    if not isinstance(obj, dict):
        return out

    items_obj = obj.get("items", [])
    if not isinstance(items_obj, list):
        return out

    for it in items_obj:
        if not isinstance(it, dict):
            continue

        item_id = str(it.get("id") or "").strip()
        justification = str(it.get("justification") or "").strip()

        if not item_id or item_id not in expected_set:
            continue
        if not justification:
            continue

        out[item_id] = justification

    return out


def _missing_expected_ids(out: Dict[str, str], expected_ids: List[str]) -> List[str]:
    return [item_id for item_id in expected_ids if not (out.get(item_id, "") or "").strip()]


def _extra_returned_ids(out: Dict[str, str], expected_ids: List[str]) -> List[str]:
    expected_set = set(expected_ids)
    return [item_id for item_id in out if item_id not in expected_set]


def generate_justifications_via_openai(batch_ctx: List[Dict[str, Any]]) -> Dict[str, str]:
    client = openai_client()
    if client is None or BaseModel is None:
        print(
            "[AI] AI runtime unavailable for justifications; deterministic fallback will be used.",
            flush=True,
        )
        return {}

    model = os.getenv("OPENAI_MODEL", "gpt-5.4").strip() or "gpt-5.4"
    effort = os.getenv("OPENAI_REASONING_EFFORT", "medium").strip() or "medium"
    configured_max_tokens = env_int("OPENAI_MAX_OUTPUT_TOKENS", 2000)
    local_endpoint = is_local_openai_compatible_endpoint()

    # For local 4B-class models, extremely large generation budgets can make JSON
    # completion less predictable. Keep the budget proportional to the requested
    # items while still honoring the configured upper bound.
    if local_endpoint:
        max_tokens = min(configured_max_tokens, max(1800, len(batch_ctx) * 500))
    else:
        max_tokens = configured_max_tokens

    supports_parse = should_use_structured_parse(client)

    class JustificationItem(BaseModel):
        id: str
        justification: str

    class JustificationBatch(BaseModel):
        items: List[JustificationItem]

    contract = get_prompt_contract("P-AIX-001")
    system = str(contract.get("system_prompt_transcript") or contract.get("system_prompt") or "")
    if not system:
        raise RuntimeError("P-AIX-001 system prompt is empty in scripts/prompt_contracts.json")

    expected_schema = contract.get("required_output_schema_object") or {
        "items": [
            {
                "id": "<exact requirement PUID>",
                "justification": "<one short English sentence>",
            }
        ]
    }
    prompt_id = str(contract.get("prompt_id") or "P-AIX-001")
    prompt_name = str(contract.get("prompt_name") or "Requirement justification")
    prompt_scope = str(contract.get("prompt_scope") or "audit_matrix")
    prompt_category = str(contract.get("prompt_category") or "primary_audit_prompt")

    original_ids = [str(item.get("id") or "") for item in batch_ctx if isinstance(item, dict)]
    final_out: Dict[str, str] = {}
    pending_ctx: List[Dict[str, Any]] = list(batch_ctx)
    last_err: Optional[Exception] = None
    call_started_at = time.time()
    attempts_used = 0

    print(
        f"[AI] Justification request prepared: items={len(batch_ctx)} "
        f"| ids={','.join(original_ids[:5])}{'...' if len(original_ids) > 5 else ''} "
        f"| model={model} | max_tokens={max_tokens} | parse={supports_parse} "
        f"| local_endpoint={local_endpoint}",
        flush=True,
    )

    for attempt in range(1, 4):
        attempts_used = attempt
        attempt_started_at = time.time()
        pending_ids = [str(item.get("id") or "") for item in pending_ctx if isinstance(item, dict)]
        user_payload = {
            "expected_count": len(pending_ids),
            "expected_ids": pending_ids,
            "batch": pending_ctx,
        }

        print(
            f"[AI] Justification attempt {attempt}/3 started: items={len(pending_ctx)}",
            flush=True,
        )

        try:
            if supports_parse:
                resp = client.responses.parse(
                    model=model,
                    input=[
                        {"role": "system", "content": system},
                        {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
                    ],
                    text_format=JustificationBatch,
                    max_output_tokens=max_tokens,
                    reasoning={"effort": effort},
                )
                parsed = getattr(resp, "output_parsed", None)
                if parsed is None:
                    raise ValueError("Structured parser returned no parsed payload")
                attempt_out = {
                    str(it.id).strip(): str(it.justification or "").strip()
                    for it in parsed.items
                    if str(it.id or "").strip() in set(pending_ids) and str(it.justification or "").strip()
                }
                parse_route = True
                json_valid = True
                schema_valid = True
            else:
                resp = client.responses.create(
                    model=model,
                    input=[
                        {"role": "system", "content": system},
                        {"role": "user", "content": json.dumps(user_payload, ensure_ascii=False)},
                    ],
                    max_output_tokens=max_tokens,
                    reasoning={"effort": effort},
                )
                txt = (getattr(resp, "output_text", "") or "").strip()
                obj = extract_json_object_from_model_output(txt)
                attempt_out = _coerce_justification_map(obj, pending_ids)
                parse_route = False
                json_valid = True
                schema_valid = isinstance(obj.get("items", []) if isinstance(obj, dict) else None, list)

            missing = _missing_expected_ids(attempt_out, pending_ids)
            extra = _extra_returned_ids(attempt_out, pending_ids)
            complete = not missing and not extra and len(attempt_out) == len(pending_ids)

            for item_id, justification in attempt_out.items():
                if item_id in set(pending_ids):
                    final_out[item_id] = justification

            prompt_call_id = record_prompt_call(
                prompt_id=prompt_id,
                prompt_name=prompt_name,
                prompt_scope=prompt_scope,
                prompt_category=prompt_category,
                source_file="scripts/ai_security_audit_requirements_excel.py",
                source_function="generate_justifications_via_openai",
                section_name="justification",
                system_prompt=system,
                user_payload=user_payload,
                expected_schema=expected_schema,
                model=model,
                provider=os.getenv("AI_PROVIDER", ""),
                max_output_tokens=max_tokens,
                reasoning_effort=effort,
                attempt_count=attempts_used,
                retry_count=max(0, attempts_used - 1),
                expected_items=len(pending_ids),
                received_items=len(attempt_out),
                json_valid=json_valid,
                schema_valid=schema_valid and complete,
                traceability_ok=complete,
                elapsed_s=time.time() - call_started_at,
                error="" if complete else f"partial response: missing={missing}; extra={extra}",
            )
            _record_llm_call(
                call_type="justification",
                expected_items=len(pending_ids),
                received_items=len(attempt_out),
                json_valid=json_valid,
                schema_valid=schema_valid and complete,
                retry_count=max(0, attempts_used - 1),
                elapsed_s=time.time() - call_started_at,
                model=model,
                max_tokens=max_tokens,
                parse_route=parse_route,
                prompt_call_id=prompt_call_id,
                prompt_id=prompt_id,
                prompt_scope=prompt_scope,
                prompt_category=prompt_category,
                error="" if complete else f"partial response: missing={missing}; extra={extra}",
            )

            if complete:
                print(
                    f"[AI] Justification attempt {attempt}/3 succeeded: "
                    f"received={len(attempt_out)}/{len(pending_ids)} "
                    f"| elapsed={time.time() - attempt_started_at:.1f}s",
                    flush=True,
                )
                break

            print(
                f"[WARN] Justification attempt {attempt}/3 produced partial JSON: "
                f"received={len(attempt_out)}/{len(pending_ids)} "
                f"| missing={len(missing)} | extra={len(extra)} "
                f"| retrying only missing items "
                f"| elapsed={time.time() - attempt_started_at:.1f}s",
                file=sys.stderr,
                flush=True,
            )

            missing_set = set(missing)
            pending_ctx = [
                item for item in pending_ctx
                if str(item.get("id") or "") in missing_set
            ]

            if not pending_ctx:
                break

        except Exception as e:
            last_err = e
            print(
                f"[WARN] Justification attempt {attempt}/3 failed: {e} "
                f"| elapsed={time.time() - attempt_started_at:.1f}s",
                file=sys.stderr,
                flush=True,
            )
            time.sleep(0.6 * attempt)

    missing_final = _missing_expected_ids(final_out, original_ids)

    if missing_final:
        print(
            f"[WARN] AI justification generation incomplete after retries: "
            f"received={len(final_out)}/{len(original_ids)} "
            f"| missing={missing_final}; deterministic fallback will be used for missing items.",
            file=sys.stderr,
            flush=True,
        )

        prompt_call_id = record_prompt_call(
            prompt_id=prompt_id,
            prompt_name=prompt_name,
            prompt_scope=prompt_scope,
            prompt_category=prompt_category,
            source_file="scripts/ai_security_audit_requirements_excel.py",
            source_function="generate_justifications_via_openai",
            section_name="justification",
            system_prompt=system,
            user_payload={"expected_count": len(original_ids), "expected_ids": original_ids, "batch": batch_ctx},
            expected_schema=expected_schema,
            model=model,
            provider=os.getenv("AI_PROVIDER", ""),
            max_output_tokens=max_tokens,
            reasoning_effort=effort,
            attempt_count=attempts_used,
            retry_count=max(0, attempts_used - 1),
            expected_items=len(original_ids),
            received_items=len(final_out),
            json_valid=False if last_err else True,
            schema_valid=False,
            traceability_ok=False,
            fallback_used=True,
            elapsed_s=time.time() - call_started_at,
            error=str(last_err or f"missing justifications: {missing_final}"),
        )
        _record_llm_call(
            call_type="justification",
            expected_items=len(original_ids),
            received_items=len(final_out),
            json_valid=False if last_err else True,
            schema_valid=False,
            retry_count=max(0, attempts_used - 1),
            elapsed_s=time.time() - call_started_at,
            model=model,
            max_tokens=max_tokens,
            parse_route=supports_parse,
            error=str(last_err or f"missing justifications: {missing_final}"),
            prompt_call_id=prompt_call_id,
            prompt_id=prompt_id,
            prompt_scope=prompt_scope,
            prompt_category=prompt_category,
        )

    return {item_id: final_out[item_id] for item_id in original_ids if item_id in final_out}


def deterministic_justification(req: RequirementAudit, flag_evidences: List[FlagEvidence], meta: Dict[str, Any]) -> str:
    def _note_hint(fe: FlagEvidence) -> str:
        return " (Fallback verdict)" if "fallback verdict" in (fe.notes or "").lower() else ""

    def _brief(fe: FlagEvidence) -> str:
        exp = f", expected={fe.expected}" if fe.expected in ("YES", "NO") else ""
        return f"{fe.id}={fe.summary_norm} (state={fe.state}, evidence_count={fe.evidence_count}{_note_hint(fe)}{exp})"

    result = req.result
    present = [fe for fe in flag_evidences if fe.outcome != "MISSING"]
    missing = [fe for fe in flag_evidences if fe.outcome == "MISSING"]
    contradicts = [fe for fe in present if fe.outcome == "CONTRADICT"]
    supports = [fe for fe in present if fe.outcome == "SUPPORT"]
    unknowns = [fe for fe in present if fe.outcome == "UNKNOWN"]

    s1 = f"Result: {result} for requirement {req.puid}."
    if result == "yes":
        s2 = "Based on the fingerprint and the flags mapped to this requirement, the application is compliant because the evaluated signals support the expected secure posture and no contradicting signals were observed among the mapped flags."
    elif result == "no":
        s2 = "Based on the fingerprint and the flags mapped to this requirement, the application is not compliant because at least one mapped signal contradicts the expected outcome."
    else:
        s2 = "Based on the fingerprint and the flags mapped to this requirement, the result is n/a because the available signals are insufficient, not applicable under the detected scenario, or contain unknown/NA coverage for the mapped flags."

    if result == "no" and contradicts:
        s3 = f"Contradicting signals: {'; '.join(_brief(fe) for fe in contradicts[:3])}."
    elif supports:
        s3 = f"Key supporting signals: {'; '.join(_brief(fe) for fe in supports[:3])}."
    else:
        any_present = present[:3]
        s3 = f"Observed mapped signals: {'; '.join(_brief(fe) for fe in any_present) if any_present else 'none'}."

    gaps = []
    if missing:
        gaps.append(f"{len(missing)} flag(s) were not present in the fingerprint (flag not present in fingerprint).")
    if unknowns:
        gaps.append("Some mapped flag summaries were NA/unknown, limiting certainty to fingerprint coverage.")
    if not gaps:
        gaps.append("This justification is derived solely from the fingerprint summaries, states, notes, and evidence counts for the mapped flags.")
    sentences = [s1, s2, s3, " ".join(gaps)]
    if meta.get("override_used") and meta.get("override_scenario_activated") is not None:
        sentences.append("Feature-gating (override scope) was applied; scenario_activated=%s based on the gate flags included in the requirement flag list." % ("TRUE" if meta.get("override_scenario_activated") else "FALSE"))
    if meta.get("conditional") and meta.get("conditional_scenario_activated") is not None:
        sentences.append("Conditional applicability was evaluated using the mapped applicability signals; scenario_activated=%s." % ("TRUE" if meta.get("conditional_scenario_activated") else "FALSE"))
    return " ".join(sentences[:6])


def main() -> None:
    run_started_at = time.time()
    RUN_METRICS["llm_calls"] = []
    RUN_METRICS["llm_items"] = []
    RUN_METRICS["adjudication_by_requirement"] = []
    RUN_METRICS["adjudication_by_flag"] = []
    strict_english = env_bool("STRICT_ENGLISH_OUTPUT", True)

    print(f"[PATH] DATA_DIR={DATA_DIR}", flush=True)
    print(f"[PATH] FINGERPRINT_PATH={FINGERPRINT_PATH}", flush=True)
    print(f"[PATH] REQUISITES_PATH={REQUISITES_PATH}", flush=True)
    print(f"[PATH] OUTPUT_XLSX_PATH={OUTPUT_XLSX_PATH}", flush=True)
    print(f"[PATH] RUN_METRICS_XLSX_PATH={RUN_METRICS_XLSX_PATH}", flush=True)

    DATA_DIR.mkdir(parents=True, exist_ok=True)
    OUTPUT_XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    RUN_METRICS_XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)

    if not FINGERPRINT_PATH.exists():
        raise SystemExit(f"Missing required file: {FINGERPRINT_PATH}.")
    if not REQUISITES_PATH.exists():
        raise SystemExit(f"Missing required file: {REQUISITES_PATH}.")

    fingerprint_data = load_json_with_one_repair(FINGERPRINT_PATH)
    requisites_data = load_json_with_one_repair(REQUISITES_PATH)
    requirements = normalize_requirements(requisites_data)
    flags_list = normalize_fingerprint_flags(fingerprint_data)

    flags_by_id: Dict[str, Dict[str, Any]] = {}
    for f in flags_list:
        if isinstance(f, dict) and f.get("id"):
            flags_by_id[str(f["id"])] = f

    raw_desc_by_puid: Dict[str, str] = {}
    non_english_puids: List[str] = []
    for req_obj in requirements:
        if not isinstance(req_obj, dict):
            continue
        puid, desc, _ = extract_req_fields(req_obj)
        if not puid:
            continue
        if not desc:
            desc = "(missing description)"
        raw_desc_by_puid[puid] = desc
        if looks_non_english(desc):
            non_english_puids.append(puid)

    if non_english_puids:
        preview = ", ".join(non_english_puids[:20])
        if len(non_english_puids) > 20:
            preview += f", ... (+{len(non_english_puids) - 20} more)"
        message = (
            "Requirement catalog language gate failed: requirement descriptions must be provided in English. "
            "The workflow no longer translates requirements with an LLM, because the catalog is a controlled input. "
            f"Non-English-like descriptions detected for PUID(s): {preview}"
        )
        if strict_english:
            raise SystemExit(message)
        print(f"[WARN] {message}", file=sys.stderr, flush=True)

    desc_en_by_puid: Dict[str, str] = dict(raw_desc_by_puid)

    batch_size = env_int("OPENAI_BATCH_SIZE", 25)
    batch_size = 25 if batch_size <= 0 else batch_size
    audits: List[RequirementAudit] = []
    counts = {"yes": 0, "no": 0, "n/a": 0}
    use_openai_just = env_bool("USE_OPENAI_JUSTIFICATIONS", False)
    total = len(requirements)
    n_batches = (total + batch_size - 1) // batch_size if total else 0

    for b in range(n_batches):
        start = b * batch_size
        end = min(total, start + batch_size)
        req_slice = requirements[start:end]
        batch_started_at = time.time()
        print(
            f"[AI] Starting batch {b + 1}/{n_batches}: "
            f"requirements {start + 1}-{end} of {total} "
            f"| batch_size={len(req_slice)} "
            f"| completed_before={len(audits)}/{total}",
            flush=True,
        )
        batch_ctx: List[Dict[str, Any]] = []
        batch_results: List[Tuple[RequirementAudit, List[FlagEvidence], Dict[str, Any]]] = []

        for req_obj in req_slice:
            if not isinstance(req_obj, dict):
                continue
            puid, _desc_raw, flag_ids = extract_req_fields(req_obj)
            if not puid:
                continue
            desc_en = desc_en_by_puid.get(puid, "(missing description)")
            result, flag_evs, meta = audit_requirement(puid, desc_en, flag_ids, flags_by_id)
            _record_adjudication_diagnostics(puid=puid, result=result, flag_ids=flag_ids, flag_evs=flag_evs, meta=meta)
            req_audit = RequirementAudit(puid=puid, description_en=desc_en, result=result, flags_used=flag_ids, justification_en="")
            batch_results.append((req_audit, flag_evs, meta))

            flags_ctx = []
            for fe in flag_evs:
                note_hint = "Fallback verdict" if "fallback verdict" in (fe.notes or "").lower() else ""
                notes_trim = (fe.notes or "")[:800]
                if strict_english and looks_non_english(notes_trim):
                    notes_trim = "Non-English notes detected in fingerprint. Do not quote verbatim; provide an English paraphrase."
                flags_ctx.append({
                    "id": fe.id,
                    "classification": fe.classification,
                    "expected": fe.expected,
                    "observed_summary_norm": fe.summary_norm,
                    "app_verdict": {"state": fe.state, "summary": fe.summary, "notes": notes_trim, "evidence_count": fe.evidence_count, "note_hint": note_hint},
                    "outcome": fe.outcome,
                })
            batch_ctx.append({"id": puid, "description_en": desc_en, "result": result, "flags_used": flag_ids, "meta": meta, "flags": flags_ctx})

        if use_openai_just:
            print(
                f"[AI] Calling local/cloud AI for batch {b + 1}/{n_batches}: "
                f"{len(batch_ctx)} requirement(s) | range={start + 1}-{end}",
                flush=True,
            )
            ai_call_started_at = time.time()
            just_map = generate_justifications_via_openai(batch_ctx)
            ai_call_elapsed_s = time.time() - ai_call_started_at
            print(
                f"[AI] AI call finished for batch {b + 1}/{n_batches}: "
                f"received={len(just_map)}/{len(batch_ctx)} justification(s) "
                f"| elapsed={ai_call_elapsed_s:.1f}s",
                flush=True,
            )
        else:
            just_map = {}
            print(
                f"[AI] AI justifications disabled for batch {b + 1}/{n_batches}; "
                f"using deterministic justifications.",
                flush=True,
            )

        missing_ai_justifications = 0
        if use_openai_just:
            missing_ai_justifications = sum(
                1 for req_audit, _flag_evs, _meta in batch_results
                if not (just_map.get(req_audit.puid, "") or "").strip()
            )

        print(
            f"[AI] Finalizing batch {b + 1}/{n_batches}: "
            f"deterministic_fallbacks={missing_ai_justifications}/{len(batch_results)}",
            flush=True,
        )

        for req_audit, flag_evs, meta in batch_results:
            ai_justification = (just_map.get(req_audit.puid, "") or "").strip()
            used_fallback = not bool(ai_justification)
            just = ai_justification or deterministic_justification(req_audit, flag_evs, meta)
            if strict_english and looks_non_english(just):
                used_fallback = True
                just = deterministic_justification(req_audit, flag_evs, meta)
                if strict_english and looks_non_english(just):
                    raise SystemExit(f"STRICT_ENGLISH_OUTPUT is enabled, but the generated justification for PUID={req_audit.puid} is not English.")
            req_audit.justification_en = just
            _append_llm_item({
                "run_batch": b + 1,
                "puid": req_audit.puid,
                "expected_by_llm": bool(use_openai_just),
                "received_from_llm": bool(ai_justification),
                "field_complete": bool(ai_justification),
                "traceability_ok": bool(ai_justification and req_audit.puid in just_map),
                "fallback_used": bool(used_fallback),
                "justification_source": "ai" if ai_justification and not used_fallback else "deterministic_fallback",
                "result": req_audit.result,
                "flags_count": len(req_audit.flags_used),
            })
            audits.append(req_audit)
            counts[req_audit.result] += 1

        batch_elapsed_s = time.time() - batch_started_at
        print(
            f"[AI] Completed batch {b + 1}/{n_batches}: "
            f"processed={len(audits)}/{total} "
            f"| yes={counts['yes']} no={counts['no']} n/a={counts['n/a']} "
            f"| elapsed={batch_elapsed_s:.1f}s",
            flush=True,
        )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "audit"
    ws.append(["id (PUID)", "Description (EN)", "Result", "Justification (EN)", "Flags used"])
    for a in audits:
        ws.append([a.puid, a.description_en, a.result, a.justification_en, ", ".join(a.flags_used)])

    OUTPUT_XLSX_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_XLSX_PATH)
    print(f"[OK] Excel generated: {OUTPUT_XLSX_PATH}", flush=True)
    _write_run_metrics_raw_xlsx(
        audits=audits,
        counts=counts,
        run_started_at=run_started_at,
        batch_size=batch_size,
        total_requirements=total,
        strict_english=strict_english,
        use_openai_justifications=use_openai_just,
        non_english_requirements_detected=len(non_english_puids),
    )
    print(f"[SUMMARY] total={len(audits)} yes={counts['yes']} no={counts['no']} n/a={counts['n/a']}", flush=True)


if __name__ == "__main__":
    try:
        main()
    except json.JSONDecodeError as e:
        print(f"JSON parsing error: {_json_decode_error_details(e)}", file=sys.stderr)
        raise SystemExit(1)
    except Exception as e:
        print(f"Error: {e}", file=sys.stderr)
        raise SystemExit(1)
    