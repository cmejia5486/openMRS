#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build stability-analysis.xlsx from multiple run-metrics packages.

Input layout supported:
  runs/run_01/run-metrics.xlsx
  runs/run_02/run-metrics.xlsx
  ...

The script can also receive explicit workbook paths through --inputs. It reads
only exported run-metrics workbooks and does not call any LLM.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import math
import statistics
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

THEME_NAVY = "17365D"
THEME_LIGHT = "F7F9FC"
THEME_BLUE = "D9EAF7"
THEME_RED = "FCE4D6"
THEME_GREEN = "E2F0D9"
TEXT = "1F2933"


def _now() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _sha256_text(value: str) -> str:
    return hashlib.sha256(str(value or "").encode("utf-8", errors="replace")).hexdigest()


def _stable_hash(value: Any) -> str:
    try:
        payload = json.dumps(value, ensure_ascii=False, sort_keys=True, default=str)
    except Exception:
        payload = str(value)
    return _sha256_text(payload)


def _safe_float(v: Any, default: float = 0.0) -> float:
    try:
        if v in (None, "", "not_applicable"):
            return default
        return float(str(v).strip())
    except Exception:
        return default


def _safe_int(v: Any, default: int = 0) -> int:
    try:
        if v in (None, ""):
            return default
        return int(float(str(v).strip()))
    except Exception:
        return default


def _read_rows(wb: Any, sheet: str) -> List[Dict[str, Any]]:
    if sheet not in wb.sheetnames:
        return []
    ws = wb[sheet]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    out: List[Dict[str, Any]] = []
    for row in rows[1:]:
        obj: Dict[str, Any] = {}
        for i, h in enumerate(headers):
            if h:
                obj[h] = row[i] if i < len(row) else None
        if any(v not in (None, "") for v in obj.values()):
            out.append(obj)
    return out


def _read_kv(wb: Any, sheet: str) -> Dict[str, Any]:
    rows = _read_rows(wb, sheet)
    return {str(r.get("key") or ""): r.get("value") for r in rows if r.get("key") not in (None, "")}


def _find_workbooks(input_dir: str, inputs: List[str]) -> List[Path]:
    paths: List[Path] = []
    for raw in inputs:
        p = Path(raw).expanduser()
        if p.is_file():
            paths.append(p)
    if input_dir:
        root = Path(input_dir).expanduser()
        if root.is_file() and root.name.lower().endswith(".xlsx"):
            paths.append(root)
        elif root.is_dir():
            paths.extend(sorted(root.glob("*/run-metrics.xlsx")))
            paths.extend(sorted(root.glob("**/run-metrics.xlsx")))
    seen = set()
    unique: List[Path] = []
    for p in paths:
        rp = str(p.resolve())
        if rp not in seen:
            seen.add(rp)
            unique.append(p)
    return unique


def _result_col(row: Dict[str, Any]) -> str:
    for key in ("result", "Result", "status", "Status"):
        if key in row:
            return str(row.get(key) or "").strip().lower()
    return ""


def _puid_col(row: Dict[str, Any]) -> str:
    for key in ("puid", "PUID", "id", "ID"):
        if key in row:
            return str(row.get(key) or "").strip()
    return ""


def _load_run(path: Path, index: int) -> Dict[str, Any]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    summary = _read_kv(wb, "run_summary")
    if not summary:
        summary = _read_kv(wb, "Run_Summary")
    input_rows = _read_rows(wb, "input_evidence") or _read_rows(wb, "input_hashes")
    compliance = _read_rows(wb, "compliance_export") or _read_rows(wb, "Appendix_Requirement_Rows")
    prompt_contract_summary = _read_rows(wb, "prompt_contract_summary") or _read_rows(wb, "Appendix_Prompt_Contract_Summary")
    metric_rows = _read_rows(wb, "llm_validation_summary")
    prompt_run = _read_rows(wb, "prompt_run_results") or _read_rows(wb, "prompt_success_summary")
    metrics: Dict[str, Any] = {}
    for r in metric_rows:
        m = str(r.get("metric") or "").strip()
        if m:
            metrics[m] = r.get("rate") if r.get("rate") not in (None, "") else r.get("value")
    for r in prompt_contract_summary:
        m = str(r.get("metric") or "").strip()
        if m:
            metrics[m] = r.get("value")
    results_by_puid = {_puid_col(r): _result_col(r) for r in compliance if _puid_col(r)}
    row_hashes = { _puid_col(r): str(r.get("row_hash") or r.get("Row Hash") or "") for r in compliance if _puid_col(r)}
    counts = Counter(v for v in results_by_puid.values() if v)
    input_fingerprint = _stable_hash([{k: r.get(k, "") for k in sorted(r)} for r in input_rows if str(r.get("component_group") or r.get("Group") or "") not in {"run_metrics"}])
    run_id = str(summary.get("run_id") or summary.get("Run ID") or path.parent.name or f"run_{index:03d}")
    return {
        "run_index": index,
        "run_id": run_id,
        "path": str(path),
        "repository": str(summary.get("repository") or ""),
        "commit_sha": str(summary.get("commit_sha") or ""),
        "workflow": str(summary.get("workflow") or ""),
        "generated_at_utc": str(summary.get("generated_at_utc") or summary.get("package_generated_at_utc") or ""),
        "software_id": str(summary.get("software_id") or summary.get("app_id") or summary.get("config.APP_NAME") or ""),
        "num_requirements": len(results_by_puid) or _safe_int(summary.get("num_requirements"), 0),
        "num_yes": counts.get("yes", _safe_int(summary.get("num_yes"), 0)),
        "num_no": counts.get("no", _safe_int(summary.get("num_no"), 0)),
        "num_na": counts.get("n/a", counts.get("na", _safe_int(summary.get("num_na"), 0))),
        "contradictory_requirement_count": _safe_int(summary.get("contradictory_requirement_count"), 0),
        "contradictory_flag_reference_count": _safe_int(summary.get("contradictory_flag_reference_count"), 0),
        "risk_precedence_count": _safe_int(summary.get("risk_precedence_count"), 0),
        "partial_evidence_requirement_count": _safe_int(summary.get("partial_evidence_requirement_count"), 0),
        "unknown_flag_reference_count": _safe_int(summary.get("unknown_flag_reference_count"), 0),
        "missing_flag_reference_count": _safe_int(summary.get("missing_flag_reference_count"), 0),
        "llm_config_hash": str(summary.get("llm_config_hash") or ""),
        "prompt_inventory_hash": str(summary.get("prompt_inventory_hash") or metrics.get("prompt_inventory_hash") or ""),
        "compliance_matrix_hash": str(summary.get("compliance_matrix_hash") or _stable_hash(row_hashes)),
        "input_fingerprint": input_fingerprint,
        "input_rows": input_rows,
        "results_by_puid": results_by_puid,
        "row_hashes": row_hashes,
        "metrics": metrics,
        "prompt_run": prompt_run,
    }


def _mode_value(values: List[str]) -> str:
    vals = [v for v in values if v not in (None, "")]
    if not vals:
        return ""
    return Counter(vals).most_common(1)[0][0]


def _comparability(runs: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], Dict[str, str]]:
    reference = {
        "repository": _mode_value([r["repository"] for r in runs]),
        "commit_sha": _mode_value([r["commit_sha"] for r in runs]),
        "input_fingerprint": _mode_value([r["input_fingerprint"] for r in runs]),
        "prompt_inventory_hash": _mode_value([r["prompt_inventory_hash"] for r in runs]),
        "llm_config_hash": _mode_value([r["llm_config_hash"] for r in runs]),
        "num_requirements": str(max([r["num_requirements"] for r in runs] or [0])),
        "puid_set_hash": _mode_value([_stable_hash(sorted(r["results_by_puid"].keys())) for r in runs]),
    }
    comparable: List[Dict[str, Any]] = []
    excluded: List[Dict[str, Any]] = []
    for r in runs:
        reasons: List[str] = []
        for key in ("repository", "commit_sha", "input_fingerprint", "prompt_inventory_hash", "llm_config_hash"):
            if reference.get(key) and str(r.get(key) or "") != reference[key]:
                reasons.append(f"{key} differs")
        if str(r.get("num_requirements") or "") != reference["num_requirements"]:
            reasons.append("num_requirements differs")
        if _stable_hash(sorted(r["results_by_puid"].keys())) != reference["puid_set_hash"]:
            reasons.append("PUID set differs")
        row = {
            "run_id": r["run_id"],
            "path": r["path"],
            "is_comparable": "false" if reasons else "true",
            "exclusion_reason": "; ".join(reasons),
            "repository": r["repository"],
            "commit_sha": r["commit_sha"],
            "input_fingerprint": r["input_fingerprint"],
            "prompt_inventory_hash": r["prompt_inventory_hash"],
            "llm_config_hash": r["llm_config_hash"],
            "num_requirements": r["num_requirements"],
        }
        if reasons:
            excluded.append(row)
        else:
            comparable.append(r)
    return comparable, excluded, reference


def _mean(values: List[float]) -> float:
    return round(statistics.mean(values), 6) if values else 0.0


def _sd(values: List[float]) -> float:
    return round(statistics.stdev(values), 6) if len(values) >= 2 else 0.0


def _pairwise_agreement(values: List[str]) -> Tuple[int, int]:
    total = 0
    agree = 0
    for i in range(len(values)):
        for j in range(i + 1, len(values)):
            if values[i] and values[j]:
                total += 1
                if values[i] == values[j]:
                    agree += 1
    return agree, total


def _fleiss_kappa(rows: List[List[str]], categories: List[str]) -> str:
    # Standard Fleiss kappa for fixed number of raters per item. If rows have a
    # variable number of non-empty values, return not_applicable.
    clean = [[v for v in row if v] for row in rows]
    if not clean:
        return "not_applicable"
    n = len(clean[0])
    if n < 2 or any(len(row) != n for row in clean):
        return "not_applicable"
    N = len(clean)
    cats = categories
    p_j = []
    for cat in cats:
        p_j.append(sum(row.count(cat) for row in clean) / (N * n))
    P_i = []
    for row in clean:
        counts = [row.count(cat) for cat in cats]
        P_i.append((sum(c * c for c in counts) - n) / (n * (n - 1)))
    P_bar = statistics.mean(P_i)
    P_e = sum(p * p for p in p_j)
    if math.isclose(1 - P_e, 0):
        return "not_applicable"
    return round((P_bar - P_e) / (1 - P_e), 6)


def _build_metrics(comparable: List[Dict[str, Any]], all_runs: List[Dict[str, Any]], excluded: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]]]:
    run_results = []
    for r in all_runs:
        m = r["metrics"]
        run_results.append({
            "run_id": r["run_id"],
            "is_comparable": "true" if r in comparable else "false",
            "num_requirements": r["num_requirements"],
            "num_yes": r["num_yes"],
            "num_no": r["num_no"],
            "num_na": r["num_na"],
            "compliance_matrix_hash": r["compliance_matrix_hash"],
            "contradictory_requirement_count": r.get("contradictory_requirement_count", 0),
            "contradictory_flag_reference_count": r.get("contradictory_flag_reference_count", 0),
            "risk_precedence_count": r.get("risk_precedence_count", 0),
            "partial_evidence_requirement_count": r.get("partial_evidence_requirement_count", 0),
            "json_valid_rate": m.get("json_valid_rate", ""),
            "schema_valid_rate": m.get("schema_valid_rate", ""),
            "completion_rate": m.get("completion_rate", ""),
            "traceability_ok_rate": m.get("traceability_ok_rate", ""),
            "fallback_rate": m.get("fallback_rate", ""),
            "retry_rate": m.get("retry_rate", ""),
            "prompt_success_rate": m.get("prompt_success_rate", ""),
        })
    all_puids = sorted(set().union(*(set(r["results_by_puid"].keys()) for r in comparable))) if comparable else []
    matrix_rows = []
    change_rows = []
    pair_agree_total = 0
    pair_total = 0
    fleiss_input = []
    for puid in all_puids:
        values = [r["results_by_puid"].get(puid, "") for r in comparable]
        value_set = sorted(set(v for v in values if v))
        agree, total = _pairwise_agreement(values)
        pair_agree_total += agree
        pair_total += total
        row = {"puid": puid}
        for r, v in zip(comparable, values):
            row[r["run_id"]] = v
        row["stable"] = "true" if len(value_set) <= 1 else "false"
        row["observed_results"] = ", ".join(value_set)
        matrix_rows.append(row)
        fleiss_input.append(values)
        if len(value_set) > 1:
            change_rows.append({"puid": puid, "observed_results": ", ".join(value_set), "run_values": json.dumps({r["run_id"]: v for r, v in zip(comparable, values)}, ensure_ascii=False)})
    matrices = [r["compliance_matrix_hash"] for r in comparable]
    most_common_matrix_count = Counter(matrices).most_common(1)[0][1] if matrices else 0
    validation_keys = ["json_valid_rate", "schema_valid_rate", "completion_rate", "traceability_ok_rate", "fallback_rate", "retry_rate", "prompt_success_rate"]
    metrics_rows = [
        {"metric": "total_run_count", "value": len(all_runs), "formula": "COUNT(all discovered run-metrics workbooks)", "interpretation": "Total workbooks supplied to the script."},
        {"metric": "comparable_run_count", "value": len(comparable), "formula": "COUNT(runs passing comparability checks)", "interpretation": "Runs included in stability calculations."},
        {"metric": "excluded_run_count", "value": len(excluded), "formula": "COUNT(runs failing comparability checks)", "interpretation": "Runs retained for audit trail but excluded from stability metrics."},
        {"metric": "exact_matrix_agreement_rate", "value": round(most_common_matrix_count / len(comparable), 6) if comparable else 0.0, "formula": "Most frequent compliance_matrix_hash count / comparable_run_count", "interpretation": "Share of comparable runs with the same complete result matrix."},
        {"metric": "requirement_result_agreement_rate", "value": round(pair_agree_total / pair_total, 6) if pair_total else 0.0, "formula": "Matching pairwise PUID results / all pairwise PUID comparisons", "interpretation": "Agreement of yes/no/n/a at requirement level."},
        {"metric": "changed_requirement_count", "value": len(change_rows), "formula": "COUNT(PUIDs with more than one observed result)", "interpretation": "Number of unstable requirements across comparable runs."},
        {"metric": "yes_count_mean", "value": _mean([r["num_yes"] for r in comparable]), "formula": "MEAN(num_yes)", "interpretation": "Average yes count across comparable runs."},
        {"metric": "yes_count_sd", "value": _sd([r["num_yes"] for r in comparable]), "formula": "STDEV.S(num_yes)", "interpretation": "Variation in yes counts across comparable runs."},
        {"metric": "no_count_mean", "value": _mean([r["num_no"] for r in comparable]), "formula": "MEAN(num_no)", "interpretation": "Average no count across comparable runs."},
        {"metric": "no_count_sd", "value": _sd([r["num_no"] for r in comparable]), "formula": "STDEV.S(num_no)", "interpretation": "Variation in no counts across comparable runs."},
        {"metric": "na_count_mean", "value": _mean([r["num_na"] for r in comparable]), "formula": "MEAN(num_na)", "interpretation": "Average n/a count across comparable runs."},
        {"metric": "na_count_sd", "value": _sd([r["num_na"] for r in comparable]), "formula": "STDEV.S(num_na)", "interpretation": "Variation in n/a counts across comparable runs."},
        {"metric": "contradictory_requirement_count_mean", "value": _mean([r.get("contradictory_requirement_count", 0) for r in comparable]), "formula": "MEAN(contradictory_requirement_count)", "interpretation": "Average number of requirements affected by contradictory mapped signals."},
        {"metric": "contradictory_requirement_count_sd", "value": _sd([r.get("contradictory_requirement_count", 0) for r in comparable]), "formula": "STDEV.S(contradictory_requirement_count)", "interpretation": "Variation in contradiction counts across comparable runs; target is 0 variation under identical inputs."},
        {"metric": "contradictory_flag_reference_count_mean", "value": _mean([r.get("contradictory_flag_reference_count", 0) for r in comparable]), "formula": "MEAN(contradictory_flag_reference_count)", "interpretation": "Average contradictory flag-reference count across comparable runs."},
        {"metric": "contradictory_flag_reference_count_sd", "value": _sd([r.get("contradictory_flag_reference_count", 0) for r in comparable]), "formula": "STDEV.S(contradictory_flag_reference_count)", "interpretation": "Variation in contradictory flag-reference counts across comparable runs."},
        {"metric": "risk_precedence_count_mean", "value": _mean([r.get("risk_precedence_count", 0) for r in comparable]), "formula": "MEAN(risk_precedence_count)", "interpretation": "Average number of decisions where risk precedence was applied."},
        {"metric": "risk_precedence_count_sd", "value": _sd([r.get("risk_precedence_count", 0) for r in comparable]), "formula": "STDEV.S(risk_precedence_count)", "interpretation": "Variation in risk-precedence decisions across comparable runs."},
        {"metric": "partial_evidence_requirement_count_mean", "value": _mean([r.get("partial_evidence_requirement_count", 0) for r in comparable]), "formula": "MEAN(partial_evidence_requirement_count)", "interpretation": "Average number of requirements with partial or unknown evidence."},
        {"metric": "partial_evidence_requirement_count_sd", "value": _sd([r.get("partial_evidence_requirement_count", 0) for r in comparable]), "formula": "STDEV.S(partial_evidence_requirement_count)", "interpretation": "Variation in partial-evidence counts across comparable runs."},
        {"metric": "prompt_inventory_stability", "value": round(Counter([r["prompt_inventory_hash"] for r in comparable]).most_common(1)[0][1] / len(comparable), 6) if comparable else 0.0, "formula": "Most frequent prompt_inventory_hash count / comparable_run_count", "interpretation": "Whether the same prompt contract set was used."},
        {"metric": "llm_config_stability", "value": round(Counter([r["llm_config_hash"] for r in comparable]).most_common(1)[0][1] / len(comparable), 6) if comparable else 0.0, "formula": "Most frequent llm_config_hash count / comparable_run_count", "interpretation": "Whether the same LLM runtime configuration was used."},
        {"metric": "fleiss_kappa", "value": _fleiss_kappa(fleiss_input, ["yes", "no", "n/a", "na"]), "formula": "Fleiss kappa over PUID categorical results", "interpretation": "Statistical agreement; not_applicable when the run count or result matrix is unsuitable."},
    ]
    for key in validation_keys:
        vals = [_safe_float(r["metrics"].get(key), 0.0) for r in comparable if r["metrics"].get(key) not in (None, "")]
        metrics_rows.append({"metric": f"{key}_mean", "value": _mean(vals), "formula": f"MEAN({key})", "interpretation": f"Average {key} across comparable runs."})
        metrics_rows.append({"metric": f"{key}_sd", "value": _sd(vals), "formula": f"STDEV.S({key})", "interpretation": f"Variation of {key} across comparable runs."})
    return run_results, matrix_rows, change_rows, metrics_rows


def _prompt_stability(comparable: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    by_prompt: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for r in comparable:
        for row in r["prompt_run"]:
            pid = str(row.get("prompt_id") or "")
            if pid:
                item = dict(row)
                item["run_id"] = r["run_id"]
                by_prompt[pid].append(item)
    out: List[Dict[str, Any]] = []
    for pid, rows in sorted(by_prompt.items()):
        call_counts = [_safe_int(r.get("call_count"), 0) for r in rows]
        success_rates = [_safe_float(r.get("prompt_success_rate"), 0.0) for r in rows if r.get("prompt_success_rate") not in (None, "", "not_applicable")]
        out.append({
            "prompt_id": pid,
            "observed_in_runs": len(rows),
            "call_count_mean": _mean(call_counts),
            "call_count_sd": _sd(call_counts),
            "prompt_success_rate_mean": _mean(success_rates),
            "prompt_success_rate_sd": _sd(success_rates),
            "interpretation": "Stable when observed_in_runs equals comparable_run_count and standard deviations remain low.",
        })
    return out


def _replace_sheet(wb: Any, name: str) -> Any:
    if name in wb.sheetnames:
        del wb[name]
    return wb.create_sheet(name)


def _write_rows(wb: Any, name: str, rows: List[Dict[str, Any]], headers: List[str]) -> None:
    ws = _replace_sheet(wb, name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def _write_kv(wb: Any, name: str, rows: List[Tuple[str, Any]]) -> None:
    ws = _replace_sheet(wb, name)
    ws.append(["key", "value"])
    for k, v in rows:
        ws.append([k, v])


def _style(wb: Any) -> None:
    header_fill = PatternFill("solid", fgColor=THEME_NAVY)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    body_font = Font(name="Arial", color=TEXT, size=9)
    side = Side(style="thin", color="D0D7DE")
    border = Border(left=side, right=side, top=side, bottom=side)
    table_names = set()
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        max_row, max_col = ws.max_row, ws.max_column
        if max_row == 0 or max_col == 0:
            continue
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for r in range(2, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
                if r % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=THEME_LIGHT)
        for c in range(1, max_col + 1):
            header = str(ws.cell(1, c).value or "").lower()
            width = 18
            if "hash" in header or "sha" in header:
                width = 36
            elif "interpretation" in header or "formula" in header or "reason" in header or "guide" in header:
                width = 46
            elif "path" in header or "run_values" in header:
                width = 44
            ws.column_dimensions[get_column_letter(c)].width = width
        if max_row >= 2 and max_col >= 2:
            try:
                base = "tbl_" + "".join(ch if ch.isalnum() else "_" for ch in ws.title.lower())[:30]
                name = base
                i = 2
                while name in table_names:
                    name = f"{base}_{i}"; i += 1
                table_names.add(name)
                tab = Table(displayName=name, ref=f"A1:{get_column_letter(max_col)}{max_row}")
                tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showColumnStripes=False)
                ws.add_table(tab)
            except Exception:
                pass


def build(args: argparse.Namespace) -> None:
    inputs = _find_workbooks(args.input_dir, args.inputs or [])
    if not inputs:
        raise SystemExit("No run-metrics.xlsx files found. Provide --input-dir or --inputs.")
    runs = [_load_run(path, idx + 1) for idx, path in enumerate(inputs)]
    comparable, excluded, reference = _comparability(runs)
    run_results, matrix_rows, change_rows, metrics_rows = _build_metrics(comparable, runs, excluded)
    prompt_rows = _prompt_stability(comparable)
    out = Path(args.output).expanduser()
    out.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    _write_kv(wb, "user_guide", [
        ("purpose", "This workbook aggregates n run-metrics workbooks from repeated executions of the same software and calculates stability metrics."),
        ("how_to_use", "Place extracted run-metrics packages in runs/run_01, runs/run_02, etc., then run: python tools/build_stability_analysis.py --input-dir runs --output stability-analysis.xlsx"),
        ("comparability_rule", "Runs are comparable only when repository, commit, critical input fingerprint, prompt inventory hash, LLM configuration hash, requirement count and PUID set match."),
        ("non_comparable_runs", "Runs that differ are not discarded. They are listed in excluded_runs with a reason."),
    ])
    metric_defs = [
        {"metric": "exact_matrix_agreement_rate", "plain_language_question": "Did the whole yes/no/n/a matrix repeat exactly?", "formula": "Most frequent compliance_matrix_hash count / comparable_run_count", "interpretation": "1.0 means all comparable runs produced the same complete matrix."},
        {"metric": "requirement_result_agreement_rate", "plain_language_question": "Did each PUID keep the same result across runs?", "formula": "Matching pairwise PUID results / all pairwise PUID comparisons", "interpretation": "Higher values indicate stable requirement-level outcomes."},
        {"metric": "changed_requirement_count", "plain_language_question": "Which requirements changed?", "formula": "COUNT(PUIDs with more than one observed result)", "interpretation": "0 means no PUID changed among comparable runs."},
        {"metric": "yes/no/n/a count SD", "plain_language_question": "Did aggregate counts change?", "formula": "STDEV.S(counts per run)", "interpretation": "0 means the aggregate count was identical across runs."},
        {"metric": "contradiction count mean and SD", "plain_language_question": "Were evidence-adjudication contradiction counters stable?", "formula": "MEAN and STDEV.S of contradictory_requirement_count and contradictory_flag_reference_count", "interpretation": "0 SD means the same contradiction profile was observed across comparable runs."},
        {"metric": "risk precedence count mean and SD", "plain_language_question": "Was the conservative risk-precedence rule applied consistently?", "formula": "MEAN and STDEV.S of risk_precedence_count", "interpretation": "0 SD means the same number of risk-precedence decisions occurred across comparable runs."},
        {"metric": "partial evidence count mean and SD", "plain_language_question": "Were partial or unknown evidence boundaries stable?", "formula": "MEAN and STDEV.S of partial_evidence_requirement_count", "interpretation": "0 SD means the same number of partial-evidence cases was observed across comparable runs."},
        {"metric": "prompt_inventory_stability", "plain_language_question": "Were the same prompt contracts used?", "formula": "Most frequent prompt_inventory_hash count / comparable_run_count", "interpretation": "1.0 means all comparable runs used the same prompt inventory."},
        {"metric": "llm_config_stability", "plain_language_question": "Was the same LLM configuration used?", "formula": "Most frequent llm_config_hash count / comparable_run_count", "interpretation": "1.0 means all comparable runs used the same LLM configuration."},
        {"metric": "validation rate mean and SD", "plain_language_question": "Were LLM controls stable?", "formula": "MEAN and STDEV.S of per-run validation rates", "interpretation": "Low SD indicates stable validation behavior."},
        {"metric": "Fleiss kappa", "plain_language_question": "What is statistical agreement across categorical results?", "formula": "Fleiss kappa over PUID yes/no/n/a results", "interpretation": "Reported when the matrix is suitable for fixed-rater categorical agreement."},
    ]
    _write_rows(wb, "metric_definitions", metric_defs, ["metric", "plain_language_question", "formula", "interpretation"])
    _write_kv(wb, "study_summary", [
        ("generated_at_utc", _now()),
        ("total_run_count", len(runs)),
        ("comparable_run_count", len(comparable)),
        ("excluded_run_count", len(excluded)),
        ("reference_repository", reference.get("repository", "")),
        ("reference_commit_sha", reference.get("commit_sha", "")),
        ("reference_prompt_inventory_hash", reference.get("prompt_inventory_hash", "")),
        ("reference_llm_config_hash", reference.get("llm_config_hash", "")),
    ])
    comp_rows = []
    for r in runs:
        ex = next((x for x in excluded if x["run_id"] == r["run_id"]), {})
        comp_rows.append({"run_id": r["run_id"], "is_comparable": "false" if ex else "true", "exclusion_reason": ex.get("exclusion_reason", ""), "repository": r["repository"], "commit_sha": r["commit_sha"], "input_fingerprint": r["input_fingerprint"], "prompt_inventory_hash": r["prompt_inventory_hash"], "llm_config_hash": r["llm_config_hash"], "num_requirements": r["num_requirements"], "path": r["path"]})
    _write_rows(wb, "run_comparability", comp_rows, ["run_id", "is_comparable", "exclusion_reason", "repository", "commit_sha", "input_fingerprint", "prompt_inventory_hash", "llm_config_hash", "num_requirements", "path"])
    _write_rows(wb, "run_results", run_results, ["run_id", "is_comparable", "num_requirements", "num_yes", "num_no", "num_na", "compliance_matrix_hash", "contradictory_requirement_count", "contradictory_flag_reference_count", "risk_precedence_count", "partial_evidence_requirement_count", "json_valid_rate", "schema_valid_rate", "completion_rate", "traceability_ok_rate", "fallback_rate", "retry_rate", "prompt_success_rate"])
    if matrix_rows:
        matrix_headers = list(matrix_rows[0].keys())
        _write_rows(wb, "requirement_result_matrix", matrix_rows, matrix_headers)
    else:
        _write_rows(wb, "requirement_result_matrix", [], ["puid", "stable", "observed_results"])
    _write_rows(wb, "requirement_changes", change_rows, ["puid", "observed_results", "run_values"])
    _write_rows(wb, "prompt_stability", prompt_rows, ["prompt_id", "observed_in_runs", "call_count_mean", "call_count_sd", "prompt_success_rate_mean", "prompt_success_rate_sd", "interpretation"])
    _write_rows(wb, "stability_metrics", metrics_rows, ["metric", "value", "formula", "interpretation"])
    _write_rows(wb, "excluded_runs", excluded, ["run_id", "path", "is_comparable", "exclusion_reason", "repository", "commit_sha", "input_fingerprint", "prompt_inventory_hash", "llm_config_hash", "num_requirements"])
    run_index = [{"run_index": r["run_index"], "run_id": r["run_id"], "path": r["path"], "generated_at_utc": r["generated_at_utc"]} for r in runs]
    _write_rows(wb, "Appendix_Run_Index", run_index, ["run_index", "run_id", "path", "generated_at_utc"])
    hash_rows = []
    for r in runs:
        for row in r["input_rows"]:
            item = dict(row)
            item["run_id"] = r["run_id"]
            hash_rows.append(item)
    hash_headers = sorted(set().union(*(row.keys() for row in hash_rows))) if hash_rows else ["run_id"]
    _write_rows(wb, "Appendix_Input_Hashes", hash_rows, hash_headers)
    _style(wb)
    wb.save(out)
    print(f"[OK] stability-analysis workbook: {out}")


def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(description="Build stability-analysis.xlsx from extracted run-metrics packages.")
    p.add_argument("--input-dir", default="", help="Directory containing run_01/run-metrics.xlsx, run_02/run-metrics.xlsx, etc.")
    p.add_argument("--inputs", nargs="*", default=[], help="Explicit run-metrics.xlsx files.")
    p.add_argument("--output", required=True, help="Output stability-analysis.xlsx path.")
    return p.parse_args()


if __name__ == "__main__":
    build(parse_args())
