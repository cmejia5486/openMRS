#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Build the user-oriented run-metrics evidence package.

The package documents one audit execution and includes the stability-analysis
builder used to aggregate n repeated executions. It never changes audit
adjudication and never invents scanner findings, prompt calls, or metrics.
"""
from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
import shutil
import sys
import textwrap
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, List, Tuple

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt

from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import Image, PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table as RLTable, TableStyle

try:
    from lib.prompt_registry import (
        prompt_contracts_map,
        registered_prompt_inventory,
        registry_path,
        registry_sha256,
        stable_json_hash,
    )
except Exception:
    def stable_json_hash(value: Any) -> str:
        try:
            payload = json.dumps(value, sort_keys=True, ensure_ascii=False, default=str)
        except Exception:
            payload = str(value)
        return hashlib.sha256(payload.encode("utf-8", errors="replace")).hexdigest()

    def registered_prompt_inventory() -> List[Dict[str, Any]]:
        return []

    def prompt_contracts_map(active_only: bool = False) -> Dict[str, Dict[str, Any]]:
        return {}

    def registry_path() -> Path:
        return Path(__file__).resolve().parent / "prompt_contracts.json"

    def registry_sha256() -> str:
        return ""

RUN_METRICS_XLSX = "run-metrics.xlsx"
RUN_METRICS_PDF = "run-metrics-methodology.pdf"
RUN_METRICS_MANIFEST = "run-metrics-manifest.json"

# Prompt contracts are centralized in scripts/prompt_contracts.json.
PROMPT_CONTRACTS_FILE_SRC = registry_path()
PROMPT_CONTRACTS_PACKAGE_PATH = "contracts/prompt_contracts.json"


def _contract_with_registry_defaults(prompt_id: str, row: Dict[str, Any]) -> Dict[str, Any]:
    out = dict(row or {})
    if prompt_id:
        out["prompt_id"] = prompt_id
    out.setdefault("registration_status", "registered")
    out.setdefault("system_prompt_transcript", out.get("system_prompt", ""))
    out.setdefault("user_payload_contract", "")
    out.setdefault("required_output_schema", "")
    out.setdefault("control_clauses", "")
    out.setdefault("validation_mechanism", "")
    if out.get("prompt_contract_hash") in (None, ""):
        out["prompt_contract_hash"] = stable_json_hash({
            "prompt_id": out.get("prompt_id", ""),
            "script_id": out.get("script_id", ""),
            "system_prompt_transcript": out.get("system_prompt_transcript", ""),
            "user_payload_contract": out.get("user_payload_contract", ""),
            "required_output_schema": out.get("required_output_schema", ""),
            "control_clauses": out.get("control_clauses", ""),
            "validation_mechanism": out.get("validation_mechanism", ""),
        })
    return out


def _combined_registered_prompt_contracts() -> Dict[str, Dict[str, Any]]:
    combined: Dict[str, Dict[str, Any]] = {}
    try:
        for row in registered_prompt_inventory(active_only=False):
            pid = str(row.get("prompt_id") or "")
            if not pid:
                continue
            combined[pid] = _contract_with_registry_defaults(pid, dict(row))
    except TypeError:
        for row in registered_prompt_inventory():
            pid = str(row.get("prompt_id") or "")
            if pid:
                combined[pid] = _contract_with_registry_defaults(pid, dict(row))
    except Exception:
        pass
    try:
        for pid, row in prompt_contracts_map(active_only=False).items():
            if pid and pid not in combined:
                combined[pid] = _contract_with_registry_defaults(pid, dict(row))
    except Exception:
        pass
    return combined


STABILITY_TOOL_SRC = Path(__file__).resolve().parent / "tools" / "build_stability_analysis.py"
STABILITY_TOOL_PACKAGE_PATH = "tools/build_stability_analysis.py"
PROMPT_REGISTRY_VALIDATOR_SRC = Path(__file__).resolve().parent / "validate_prompt_registry.py"
PROMPT_REGISTRY_VALIDATOR_PACKAGE_PATH = "tools/validate_prompt_registry.py"

THEME_NAVY = "17365D"
THEME_BLUE = "D9EAF7"
THEME_LIGHT_BLUE = "EEF5FB"
THEME_LIGHT = "F7F9FC"
THEME_TEXT = "1F2933"
THEME_GREEN = "E2F0D9"
THEME_RED = "FCE4D6"
THEME_ORANGE = "FCE4D6"
THEME_GRAY = "EDEDED"


NUMERIC_DECISION_RULES = {
    "num_requirements": "comparability requirement: value must equal the same metric in every comparable run; mismatch excludes the run from stability analysis.",
    "compliance_matrix_hash": "exact agreement requirement: value must match across comparable runs for exact_matrix_agreement_rate = 1.000.",
    "llm_config_hash": "comparability requirement: value must match across comparable runs unless intentionally comparing configurations.",
    "prompt_inventory_hash": "comparability requirement: value must match across comparable runs; target prompt_inventory_stability = 1.000.",
    "json_valid_rate": "acceptable >= 0.950; review-required >= 0.900 and < 0.950; critical < 0.900.",
    "schema_valid_rate": "acceptable >= 0.950; review-required >= 0.900 and < 0.950; critical < 0.900.",
    "completion_rate": "acceptable >= 0.950; review-required >= 0.900 and < 0.950; critical < 0.900.",
    "traceability_ok_rate": "acceptable >= 0.950; review-required >= 0.900 and < 0.950; critical < 0.900.",
    "fallback_rate": "target <= 0.050; review-required > 0.050 and <= 0.100; critical > 0.100.",
    "retry_rate": "target <= 0.200; review-required > 0.200 and <= 0.500; critical > 0.500.",
    "prompt_success_rate": "acceptable >= 0.950; review-required >= 0.900 and < 0.950; critical < 0.900.",
}

STABILITY_NUMERIC_DECISION_RULES = {
    "comparable_run_count": "minimum n >= 2 for a repeated-run comparison; recommended n >= 5 when the study protocol requires five executions.",
    "exact_matrix_agreement_rate": "target = 1.000; review-required >= 0.950 and < 1.000; critical < 0.950.",
    "requirement_result_agreement_rate": "acceptable >= 0.990; review-required >= 0.950 and < 0.990; critical < 0.950.",
    "changed_requirement_count": "target = 0; review-required 1 to 2; critical >= 3.",
    "yes_count_sd": "target = 0.000; review-required > 0.000 and <= 1.000; critical > 1.000.",
    "no_count_sd": "target = 0.000; review-required > 0.000 and <= 1.000; critical > 1.000.",
    "na_count_sd": "target = 0.000; review-required > 0.000 and <= 1.000; critical > 1.000.",
    "prompt_inventory_stability": "target = 1.000; review-required >= 0.950 and < 1.000; critical < 0.950.",
    "llm_config_stability": "target = 1.000; review-required >= 0.950 and < 1.000; critical < 0.950.",
    "validation_rate_mean": "positive rates acceptable >= 0.950; fallback mean target <= 0.050; retry mean target <= 0.200.",
    "validation_rate_sd": "target <= 0.020; review-required > 0.020 and <= 0.050; critical > 0.050.",
    "Fleiss kappa": "strong agreement >= 0.800; moderate >= 0.600 and < 0.800; weak < 0.600; not_applicable when n or category distribution is unsuitable.",
}

def _metric_decision_rule(metric: str) -> str:
    return NUMERIC_DECISION_RULES.get(str(metric or ""), "explicit numeric rule not configured; review metric definition before publication.")

CONTROL_FIELDS = [
    "json_required_detected",
    "schema_required_detected",
    "identifier_preservation_detected",
    "grounding_required_detected",
    "no_invent_evidence_detected",
    "no_change_result_detected",
    "english_only_detected",
    "repair_or_retry_detected",
]

CONTROL_LABELS = {
    "json_required_detected": "JSON-only output",
    "schema_required_detected": "Expected schema",
    "identifier_preservation_detected": "Identifier preservation",
    "grounding_required_detected": "Grounded in supplied evidence",
    "no_invent_evidence_detected": "No invented evidence",
    "no_change_result_detected": "No result modification",
    "english_only_detected": "English-only output",
    "repair_or_retry_detected": "Retry or repair path",
}


def _infer_control_name_and_metric(clause: Any) -> Tuple[str, str]:
    text = str(clause or "").lower()
    if "json" in text or "raw json" in text:
        return "JSON-only output", "json_valid_rate"
    if "precomputed result" in text or "do not change" in text or "does not adjudicate" in text:
        return "No result modification", "compliance_matrix_hash"
    if "do not invent" in text or "unprovided" in text or "no invented" in text or "do_not_invent" in text:
        return "No invented evidence", "prompt_success_rate"
    if "exact input item_id" in text or "exact item_id" in text or "puid" in text or "finding_id" in text or "exact pattern" in text or "identifier" in text:
        return "Identifier preservation", "traceability_ok_rate"
    if "schema" in text or "field" in text or "non-empty" in text or "result object" in text or "required" in text:
        return "Expected schema", "schema_valid_rate"
    if "english" in text:
        return "English-only output", "prompt_success_rate"
    if "repair" in text or "retry" in text or "missing" in text or "deterministic fallback" in text:
        return "Retry, repair, or fallback path", "retry_rate"
    if "use only" in text or "provided" in text or "supplied" in text or "evidence" in text or "observed" in text or "ground" in text:
        return "Grounded in supplied evidence", "prompt_success_rate"
    return "Registered prompt-control clause", "prompt_success_rate"

METRIC_DEFINITIONS = [
    {
        "metric": "num_requirements",
        "plain_language_question": "How many requirements were evaluated in this execution?",
        "purpose": "Confirms the evaluated requirement population.",
        "formula": "COUNTROWS(Appendix_Requirement_Rows)",
        "numerator": "Rows in Appendix_Requirement_Rows",
        "denominator": "Not applicable",
        "source_sheet": "Appendix_Requirement_Rows",
        "interpretation": "Comparable repeated runs must have identical num_requirements and identical PUID sets; otherwise the run is excluded from stability calculations.",
        "single_run": "yes",
        "stability_analysis": "yes",
    },
    {
        "metric": "compliance_matrix_hash",
        "plain_language_question": "Did the complete yes/no/n/a matrix remain exactly identifiable?",
        "purpose": "Fingerprints the ordered matrix of requirement results.",
        "formula": "SHA-256 of ordered row hashes by PUID",
        "numerator": "Not applicable",
        "denominator": "Not applicable",
        "source_sheet": "Run_Summary and Appendix_Requirement_Rows",
        "interpretation": "Exact matrix agreement is satisfied only when this hash is identical across comparable runs; target exact_matrix_agreement_rate = 1.000.",
        "single_run": "yes",
        "stability_analysis": "yes",
    },
    {
        "metric": "llm_config_hash",
        "plain_language_question": "Was the same LLM configuration used?",
        "purpose": "Fingerprints model, provider and relevant runtime settings.",
        "formula": "SHA-256 of normalized LLM configuration snapshot",
        "numerator": "Not applicable",
        "denominator": "Not applicable",
        "source_sheet": "Run_Summary",
        "interpretation": "Comparable repeated runs must use the same llm_config_hash; otherwise the run is excluded unless the study intentionally compares configurations.",
        "single_run": "yes",
        "stability_analysis": "yes",
    },
    {
        "metric": "prompt_inventory_hash",
        "plain_language_question": "Were the same prompt contracts available?",
        "purpose": "Fingerprints the registered prompt contracts and observed execution state.",
        "formula": "SHA-256 of normalized Prompt_Contracts rows and observed prompt hashes",
        "numerator": "Not applicable",
        "denominator": "Not applicable",
        "source_sheet": "Prompt_Contracts",
        "interpretation": "Comparable repeated runs must use the same prompt_inventory_hash; target prompt_inventory_stability = 1.000.",
        "single_run": "yes",
        "stability_analysis": "yes",
    },
    {
        "metric": "json_valid_rate",
        "plain_language_question": "Did the LLM return parseable JSON?",
        "purpose": "Measures compliance with the JSON-only output contract.",
        "formula": "json_valid_count / num_llm_calls",
        "numerator": "LLM calls with json_valid=true",
        "denominator": "All LLM calls recorded",
        "source_sheet": "Appendix_LLM_Calls and Appendix_Prompt_Calls",
        "interpretation": "Apply numeric rule: acceptable >= 0.950; review-required >= 0.900 and < 0.950; critical < 0.900.",
        "single_run": "yes",
        "stability_analysis": "yes",
    },
    {
        "metric": "schema_valid_rate",
        "plain_language_question": "Did the LLM return the fields expected by the workflow?",
        "purpose": "Measures structural validity of LLM output.",
        "formula": "schema_valid_count / num_llm_calls",
        "numerator": "LLM calls with schema_valid=true",
        "denominator": "All LLM calls recorded",
        "source_sheet": "Appendix_LLM_Calls and Appendix_Prompt_Calls",
        "interpretation": "Apply numeric rule: acceptable >= 0.950; review-required >= 0.900 and < 0.950; critical < 0.900.",
        "single_run": "yes",
        "stability_analysis": "yes",
    },
    {
        "metric": "completion_rate",
        "plain_language_question": "Did the LLM return the expected number of items?",
        "purpose": "Checks whether requested PUIDs or item_ids were returned.",
        "formula": "received_items / expected_items",
        "numerator": "Received items",
        "denominator": "Expected items",
        "source_sheet": "Appendix_LLM_Calls and Appendix_Prompt_Calls",
        "interpretation": "Apply numeric rule: acceptable >= 0.950; review-required >= 0.900 and < 0.950; critical < 0.900.",
        "single_run": "yes",
        "stability_analysis": "yes",
    },
    {
        "metric": "traceability_ok_rate",
        "plain_language_question": "Did the LLM preserve the expected PUID or item_id?",
        "purpose": "Prevents identifier drift in generated text.",
        "formula": "traceability_ok_items / traceability_expected_items",
        "numerator": "Items preserving the expected identifier",
        "denominator": "Items where identifier traceability applies",
        "source_sheet": "Appendix_LLM_Items and Appendix_Prompt_Calls",
        "interpretation": "Apply numeric rule: acceptable >= 0.950; review-required >= 0.900 and < 0.950; critical < 0.900.",
        "single_run": "yes",
        "stability_analysis": "yes",
    },
    {
        "metric": "fallback_rate",
        "plain_language_question": "How often did deterministic fallback replace LLM output?",
        "purpose": "Shows how often model output was absent, incomplete or rejected.",
        "formula": "fallback_items / num_requirements",
        "numerator": "Items with fallback_used=true",
        "denominator": "Requirements evaluated",
        "source_sheet": "Appendix_LLM_Items",
        "interpretation": "Apply numeric rule: target <= 0.050; review-required > 0.050 and <= 0.100; critical > 0.100.",
        "single_run": "yes",
        "stability_analysis": "yes",
    },
    {
        "metric": "retry_rate",
        "plain_language_question": "How often did the workflow retry an LLM call?",
        "purpose": "Measures operational instability or invalid response recovery.",
        "formula": "retry_count / num_llm_calls",
        "numerator": "Total retry count",
        "denominator": "All LLM calls recorded",
        "source_sheet": "Appendix_LLM_Calls and Appendix_Prompt_Calls",
        "interpretation": "Apply numeric rule: target <= 0.200; review-required > 0.200 and <= 0.500; critical > 0.500.",
        "single_run": "yes",
        "stability_analysis": "yes",
    },
    {
        "metric": "prompt_success_rate",
        "plain_language_question": "Which prompts succeeded in this run?",
        "purpose": "Summarizes acceptance at prompt-contract level.",
        "formula": "successful_prompt_calls / executed_prompt_calls",
        "numerator": "Accepted prompt calls",
        "denominator": "Executed prompt calls",
        "source_sheet": "Prompt_Run_Results and Appendix_Prompt_Calls",
        "interpretation": "Apply numeric rule: acceptable >= 0.950; review-required >= 0.900 and < 0.950; critical < 0.900.",
        "single_run": "yes",
        "stability_analysis": "yes",
    },
]


for _metric_def in METRIC_DEFINITIONS:
    _metric_def.setdefault("numeric_decision_rule", _metric_decision_rule(_metric_def.get("metric")))


METRIC_DEFINITIONS.extend([
    {
        "metric": "contradictory_requirement_count",
        "plain_language_question": "How many requirements had at least one mapped signal contradicting the expected secure posture?",
        "purpose": "Documents evidence conflicts handled by deterministic adjudication.",
        "formula": "COUNTROWS(adjudication_by_requirement where contradictory_requirement=true)",
        "numerator": "Requirements with at least one contradictory non-applicability flag",
        "denominator": "Not applicable",
        "source_sheet": "adjudication_by_requirement and adjudication_summary",
        "interpretation": "A contradiction does not indicate workflow failure; it indicates that risk evidence was present and must be handled conservatively.",
        "numeric_decision_rule": "Tracked for transparency. Any value > 0 should be reviewable through adjudication_by_requirement and adjudication_by_flag.",
        "single_run": "yes",
        "stability_analysis": "yes",
    },
    {
        "metric": "contradictory_flag_reference_count",
        "plain_language_question": "How many mapped flag references contradicted the expected posture?",
        "purpose": "Counts contradictory flag-level evidence references across all requirements.",
        "formula": "COUNTROWS(adjudication_by_flag where outcome=CONTRADICT)",
        "numerator": "Contradictory flag references",
        "denominator": "All adjudicated flag references",
        "source_sheet": "adjudication_by_flag and adjudication_summary",
        "interpretation": "Used to explain how frequently risk signals overrode positive or inconclusive signals.",
        "numeric_decision_rule": "Tracked for transparency and cross-run stability; inspect affected PUIDs when non-zero.",
        "single_run": "yes",
        "stability_analysis": "yes",
    },
    {
        "metric": "risk_precedence_count",
        "plain_language_question": "How many requirement decisions applied the conservative risk-precedence rule?",
        "purpose": "Shows when a non-compliant result was caused by mapped risk evidence.",
        "formula": "COUNTROWS(adjudication_by_requirement where risk_precedence_applied=true)",
        "numerator": "Requirements resolved by risk precedence",
        "denominator": "Not applicable",
        "source_sheet": "adjudication_by_requirement and adjudication_summary",
        "interpretation": "When risk precedence is applied, positive or absent evidence from other sources does not neutralize the mapped risk signal.",
        "numeric_decision_rule": "Tracked for auditability. Stability across repeated runs should be assessed.",
        "single_run": "yes",
        "stability_analysis": "yes",
    },
    {
        "metric": "partial_evidence_requirement_count",
        "plain_language_question": "How many requirements had partial, missing, or unknown evidence without a direct contradiction?",
        "purpose": "Documents non-conclusive evidence boundaries.",
        "formula": "COUNTROWS(adjudication_by_requirement where partial_evidence=true)",
        "numerator": "Requirements with partial or unknown evidence",
        "denominator": "Not applicable",
        "source_sheet": "adjudication_by_requirement and adjudication_summary",
        "interpretation": "These cases should not be reported as compliant unless additional evidence is collected.",
        "numeric_decision_rule": "Tracked for transparency. High counts indicate evidence gaps or need for manual review.",
        "single_run": "yes",
        "stability_analysis": "yes",
    },
])

for _metric_def in METRIC_DEFINITIONS:
    _metric_def.setdefault("numeric_decision_rule", _metric_decision_rule(_metric_def.get("metric")))

PACKAGE_COMPONENTS = [
    {
        "component": "run-metrics.xlsx",
        "type": "primary",
        "purpose": "Main workbook for one audit execution. It presents user-readable summaries first and preserves technical evidence in Appendix sheets.",
        "reader_priority": "open_first",
        "duplication_policy": "Consolidates key metrics from JSONL telemetry; raw telemetry remains in the package for reconstruction.",
    },
    {
        "component": "run-metrics-methodology.pdf",
        "type": "primary",
        "purpose": "Human-readable methodology explaining package contents, metrics, prompt controls, output validation, and repeated-run stability analysis.",
        "reader_priority": "open_first",
        "duplication_policy": "Explains the workbook and tools without duplicating all raw rows.",
    },
    {
        "component": "run-metrics-manifest.json",
        "type": "integrity_manifest",
        "purpose": "Machine-readable list of packaged files with SHA-256 hashes, sizes, repository, commit and run identifiers.",
        "reader_priority": "supporting",
        "duplication_policy": "Only file integrity metadata is duplicated into Appendix_Package_Manifest.",
    },
    {
        "component": "contracts/prompt_contracts.json",
        "type": "prompt_contract_registry",
        "purpose": "Central JSON registry containing the exact LLM prompt transcripts, unique prompt IDs, script IDs, auxiliary flags, expected output schemas, control clauses and validation mechanisms used by the workflow.",
        "reader_priority": "supporting",
        "duplication_policy": "Transcribed into prompt_contracts and prompt_control_breakdown; this JSON remains the authoritative source.",
    },
    {
        "component": "tools/validate_prompt_registry.py",
        "type": "prompt_registry_validator",
        "purpose": "Standalone Python validator that checks the prompt registry for complete transcripts, schemas, controls, validation mechanisms, and LLM-call coverage.",
        "reader_priority": "supporting",
        "duplication_policy": "Validator logic is documented in the PDF and workflow; it does not duplicate runtime data.",
    },
    {
        "component": "tools/build_stability_analysis.py",
        "type": "repeated_run_tool",
        "purpose": "Standalone Python tool that reads extracted run-metrics packages from runs/run_01, runs/run_02, ..., runs/run_n and generates stability-analysis.xlsx.",
        "reader_priority": "supporting",
        "duplication_policy": "Tool logic is documented in the PDF and metric_definitions; no output is invented inside a single run.",
    },
    {
        "component": "telemetry/prompt-telemetry-ai-requirements.jsonl",
        "type": "raw_telemetry",
        "purpose": "Raw LLM prompt-call telemetry from the requirement-level matrix generation stage.",
        "reader_priority": "appendix_raw",
        "duplication_policy": "Summarized into Prompt_Run_Results and Appendix_Prompt_Calls.",
    },
    {
        "component": "telemetry/prompt-telemetry-audit-summary.jsonl",
        "type": "raw_telemetry",
        "purpose": "Raw LLM prompt-call telemetry from the Audit Summary generation stage.",
        "reader_priority": "appendix_raw",
        "duplication_policy": "Summarized into Prompt_Run_Results and Appendix_Prompt_Calls.",
    },
]

# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _now_utc() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _sha256_file(path: Path) -> str:
    try:
        if not path.is_file():
            return ""
        h = hashlib.sha256()
        with path.open("rb") as fh:
            for chunk in iter(lambda: fh.read(1024 * 1024), b""):
                h.update(chunk)
        return h.hexdigest()
    except Exception:
        return ""


def _sha256_dir(path: Path) -> Tuple[str, int]:
    if not path.is_dir():
        return "", 0
    h = hashlib.sha256(); count = 0
    for fp in sorted([p for p in path.rglob("*") if p.is_file()], key=lambda p: str(p.relative_to(path)).lower()):
        rel = str(fp.relative_to(path)).replace("\\", "/")
        h.update(rel.encode("utf-8", errors="replace")); h.update(b"\0")
        h.update(_sha256_file(fp).encode("ascii", errors="ignore")); h.update(b"\n")
        count += 1
    return (h.hexdigest() if count else ""), count


def _artifact_hash(path: Path) -> Tuple[str, str, int]:
    if path.is_file():
        return "file", _sha256_file(path), 1
    if path.is_dir():
        sha, count = _sha256_dir(path)
        return "directory", sha, count
    return "missing", "", 0


def _safe_float(value: Any, default: float = 0.0) -> float:
    try:
        if value in (None, "", "not_applicable"):
            return default
        return float(str(value).strip())
    except Exception:
        return default


def _safe_int(value: Any, default: int = 0) -> int:
    try:
        if value in (None, ""):
            return default
        if isinstance(value, bool):
            return int(value)
        return int(float(str(value).strip()))
    except Exception:
        return default


def _to_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in {"1", "true", "yes", "y", "on", "verdadero", "si", "sí"}


def _bool_text(value: Any) -> str:
    if value in (None, ""):
        return ""
    return "true" if _to_bool(value) else "false"


def _rate(numerator: float, denominator: float) -> Any:
    if denominator <= 0:
        return "not_applicable"
    return round(float(numerator) / float(denominator), 6)


def _stable_hash(value: Any) -> str:
    return stable_json_hash(value)


def _read_rows(wb: Any, sheet_name: str) -> List[Dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h or "").strip() for h in rows[0]]
    out: List[Dict[str, Any]] = []
    for row in rows[1:]:
        item: Dict[str, Any] = {}
        for idx, header in enumerate(headers):
            if header:
                item[header] = row[idx] if idx < len(row) else None
        if any(v not in (None, "") for v in item.values()):
            out.append(item)
    return out


def _read_kv(wb: Any, sheet_name: str) -> Dict[str, Any]:
    rows = _read_rows(wb, sheet_name)
    return {str(r.get("key") or ""): r.get("value") for r in rows if r.get("key") not in (None, "")}


def _result_value(row: Dict[str, Any]) -> str:
    for key in ("result", "Result", "status", "Status"):
        if key in row:
            return str(row.get(key) or "").strip().lower()
    return ""


def _puid_value(row: Dict[str, Any]) -> str:
    for key in ("puid", "PUID", "id", "ID"):
        if key in row:
            return str(row.get(key) or "").strip()
    return ""


def _first_nonempty(*values: Any) -> str:
    for v in values:
        if v not in (None, ""):
            return str(v)
    return ""


def _short(value: Any, limit: int = 500) -> str:
    text = str(value or "").strip()
    if len(text) <= limit:
        return text
    return text[:limit].rstrip() + " [text shortened]"

# ---------------------------------------------------------------------------
# Data extraction
# ---------------------------------------------------------------------------

def _compute_base_metrics(wb: Any) -> Dict[str, Any]:
    summary = _read_kv(wb, "run_summary") or _read_kv(wb, "Run_Summary")
    compliance = _read_rows(wb, "compliance_export") or _read_rows(wb, "Appendix_Requirement_Rows")
    llm_calls = _read_rows(wb, "llm_calls") or _read_rows(wb, "Appendix_LLM_Calls")
    llm_items = _read_rows(wb, "llm_items") or _read_rows(wb, "Appendix_LLM_Items")
    adjudication_summary = _read_kv(wb, "adjudication_summary") or _read_kv(wb, "Adjudication_Summary")
    adjudication_by_requirement = _read_rows(wb, "adjudication_by_requirement") or _read_rows(wb, "Adjudication_By_Requirement")
    adjudication_by_flag = _read_rows(wb, "adjudication_by_flag") or _read_rows(wb, "Adjudication_By_Flag")
    counts = Counter(_result_value(r) for r in compliance if _result_value(r))
    num_requirements = len(compliance) or _safe_int(summary.get("num_requirements"), 0)
    json_valid_count = sum(1 for r in llm_calls if _to_bool(r.get("json_valid")))
    schema_valid_count = sum(1 for r in llm_calls if _to_bool(r.get("schema_valid")))
    retry_count = sum(_safe_int(r.get("retry_count"), 0) for r in llm_calls)
    expected_items = sum(_safe_int(r.get("expected_items"), 0) for r in llm_calls)
    received_items = sum(_safe_int(r.get("received_items"), 0) for r in llm_calls)
    traceability_expected = sum(1 for r in llm_items if str(r.get("expected_by_llm") or "").strip().lower() in {"true", "1", "yes"})
    if traceability_expected == 0:
        traceability_expected = sum(1 for r in llm_items if r.get("traceability_ok") not in (None, ""))
    traceability_ok = sum(1 for r in llm_items if _to_bool(r.get("traceability_ok")))
    fallback_items = sum(1 for r in llm_items if _to_bool(r.get("fallback_used")))
    row_hashes = []
    for r in compliance:
        puid = _puid_value(r)
        result = _result_value(r)
        flags = str(r.get("flags_used") or r.get("flags") or r.get("Flags") or "")
        rh = str(r.get("row_hash") or "") or _stable_hash({"puid": puid, "result": result, "flags": flags})
        if puid:
            row_hashes.append((puid, rh))
    compliance_matrix_hash = summary.get("compliance_matrix_hash") or _stable_hash(row_hashes)
    if not adjudication_summary and adjudication_by_requirement:
        adjudication_summary = {
            "adjudicated_requirement_count": len(adjudication_by_requirement),
            "adjudicated_flag_reference_count": len(adjudication_by_flag),
            "contradictory_requirement_count": sum(1 for r in adjudication_by_requirement if _to_bool(r.get("contradictory_requirement"))),
            "contradictory_flag_reference_count": sum(1 for r in adjudication_by_flag if str(r.get("outcome") or "") == "CONTRADICT"),
            "risk_precedence_count": sum(1 for r in adjudication_by_requirement if _to_bool(r.get("risk_precedence_applied"))),
            "partial_evidence_requirement_count": sum(1 for r in adjudication_by_requirement if _to_bool(r.get("partial_evidence"))),
            "unknown_flag_reference_count": sum(1 for r in adjudication_by_flag if str(r.get("outcome") or "") == "UNKNOWN"),
            "missing_flag_reference_count": sum(1 for r in adjudication_by_flag if str(r.get("outcome") or "") == "MISSING"),
        }
    return {
        "summary": summary,
        "compliance": compliance,
        "llm_calls": llm_calls,
        "llm_items": llm_items,
        "adjudication_summary": adjudication_summary,
        "adjudication_by_requirement": adjudication_by_requirement,
        "adjudication_by_flag": adjudication_by_flag,
        "num_requirements": num_requirements,
        "counts": {"yes": counts.get("yes", 0), "no": counts.get("no", 0), "n/a": counts.get("n/a", counts.get("na", 0))},
        "num_llm_calls": len(llm_calls),
        "json_valid_count": json_valid_count,
        "schema_valid_count": schema_valid_count,
        "retry_count": retry_count,
        "expected_items": expected_items,
        "received_items": received_items,
        "traceability_expected": traceability_expected,
        "traceability_ok": traceability_ok,
        "fallback_items": fallback_items,
        "json_valid_rate": _rate(json_valid_count, len(llm_calls)),
        "schema_valid_rate": _rate(schema_valid_count, len(llm_calls)),
        "completion_rate": _rate(received_items, expected_items),
        "traceability_ok_rate": _rate(traceability_ok, traceability_expected),
        "fallback_rate": _rate(fallback_items, num_requirements),
        "retry_rate": _rate(retry_count, len(llm_calls)),
        "compliance_matrix_hash": compliance_matrix_hash,
    }


def _copy_telemetry_files(paths: List[str], out_dir: Path) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    events: List[Dict[str, Any]] = []
    components: List[Dict[str, Any]] = []
    telemetry_dir = out_dir / "telemetry"
    telemetry_dir.mkdir(parents=True, exist_ok=True)
    for raw in paths:
        src = Path(raw).expanduser()
        if not src.is_file():
            continue
        name = src.name
        if "audit-summary" in name or "audit_summary" in name:
            dst_name = "prompt-telemetry-audit-summary.jsonl"
            stage = "audit_summary"
        elif "ai-requirements" in name or "ai_requirements" in name:
            dst_name = "prompt-telemetry-ai-requirements.jsonl"
            stage = "ai_requirements"
        else:
            dst_name = name
            stage = "prompt_telemetry"
        dst = telemetry_dir / dst_name
        if src.resolve() != dst.resolve():
            shutil.copy2(src, dst)
        line_count = 0
        with dst.open("r", encoding="utf-8", errors="replace") as fh:
            for line in fh:
                if not line.strip():
                    continue
                line_count += 1
                try:
                    obj = json.loads(line)
                    if isinstance(obj, dict):
                        obj.setdefault("telemetry_stage", stage)
                        events.append(obj)
                except Exception as exc:
                    events.append({"event_type": "telemetry_parse_error", "telemetry_stage": stage, "message": str(exc), "source_file": str(dst)})
        components.append({"component": f"telemetry/{dst_name}", "path": str(dst), "role": "raw_telemetry", "stage": stage, "event_count": line_count, "sha256": _sha256_file(dst), "size_bytes": dst.stat().st_size})
    return events, components


def _events_from_existing_workbook(wb: Any) -> List[Dict[str, Any]]:
    out = []
    for r in _read_rows(wb, "llm_calls"):
        out.append({
            "event_type": "prompt_call",
            "telemetry_stage": "ai_requirements_workbook_fallback",
            "prompt_call_id": r.get("prompt_call_id") or r.get("call_id") or "",
            "prompt_id": r.get("prompt_id") or "P-AIX-001",
            "prompt_name": "Requirement justification",
            "prompt_scope": r.get("prompt_scope") or "audit_matrix",
            "prompt_category": r.get("prompt_category") or "primary_audit_prompt",
            "section_name": r.get("call_type") or "justification",
            "source_file": "scripts/ai_security_audit_requirements_excel.py",
            "source_function": "generate_justifications_via_openai",
            "model": r.get("model") or "",
            "max_output_tokens": r.get("max_tokens") or "",
            "attempt_count": 1 + _safe_int(r.get("retry_count"), 0),
            "retry_count": _safe_int(r.get("retry_count"), 0),
            "expected_items": _safe_int(r.get("expected_items"), 0),
            "received_items": _safe_int(r.get("received_items"), 0),
            "json_valid": _to_bool(r.get("json_valid")),
            "schema_valid": _to_bool(r.get("schema_valid")),
            "traceability_ok": "",
            "repair_used": False,
            "fallback_used": False,
            "registration_status": "registered",
        })
    return out


def _prompt_call_rows(events: List[Dict[str, Any]], fallback_events: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    raw = [e for e in events if e.get("event_type") == "prompt_call"] or fallback_events
    rows = []
    for idx, e in enumerate(raw, start=1):
        trace = e.get("traceability_ok")
        trace_text = "not_applicable" if trace in (None, "") else _bool_text(trace)
        json_ok = _to_bool(e.get("json_valid"))
        schema_ok = _to_bool(e.get("schema_valid"))
        trace_ok = trace_text != "false"
        fallback = _to_bool(e.get("fallback_used"))
        success = json_ok and schema_ok and trace_ok and not fallback
        rows.append({
            "global_prompt_call_id": f"GPCALL-{idx:04d}",
            "prompt_call_id": e.get("prompt_call_id", ""),
            "telemetry_stage": e.get("telemetry_stage", ""),
            "prompt_id": e.get("prompt_id", ""),
            "prompt_name": e.get("prompt_name", ""),
            "prompt_scope": e.get("prompt_scope", ""),
            "prompt_category": e.get("prompt_category", ""),
            "section_name": e.get("section_name", ""),
            "source_file": e.get("source_file", ""),
            "source_function": e.get("source_function", ""),
            "model": e.get("model", ""),
            "provider": e.get("provider", ""),
            "max_output_tokens": e.get("max_output_tokens", ""),
            "reasoning_effort": e.get("reasoning_effort", ""),
            "attempt_count": _safe_int(e.get("attempt_count"), 0),
            "retry_count": _safe_int(e.get("retry_count"), 0),
            "expected_items": _safe_int(e.get("expected_items"), 0),
            "received_items": _safe_int(e.get("received_items"), 0),
            "json_valid": _bool_text(json_ok),
            "schema_valid": _bool_text(schema_ok),
            "traceability_ok": trace_text,
            "repair_used": _bool_text(e.get("repair_used")),
            "fallback_used": _bool_text(fallback),
            "prompt_success": "true" if success else "false",
            "prompt_hash": e.get("prompt_hash", ""),
            "schema_hash": e.get("schema_hash", ""),
            "payload_hash": e.get("payload_hash", ""),
            "controls_hash": e.get("controls_hash", ""),
            "contract_score": e.get("contract_score", ""),
            "elapsed_s": e.get("elapsed_s", ""),
            "registration_status": e.get("registration_status", "registered"),
            "error": e.get("error", ""),
        })
    return rows


def _registered_contract_rows() -> List[Dict[str, Any]]:
    return [dict(row) for _, row in sorted(_combined_registered_prompt_contracts().items())]


def _build_prompt_sections(prompt_calls: List[Dict[str, Any]]) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]], List[Dict[str, Any]], str]:
    registered = {str(r.get("prompt_id") or ""): dict(r) for r in _registered_contract_rows() if r.get("prompt_id")}
    by_prompt: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for c in prompt_calls:
        pid = str(c.get("prompt_id") or "")
        by_prompt[pid].append(c)
        if pid and pid not in registered:
            runtime_row = {"prompt_id": pid, "prompt_name": c.get("prompt_name", "Unregistered prompt"), "prompt_scope": c.get("prompt_scope", "unregistered"), "prompt_category": "unregistered_candidate", "registration_status": "unregistered", "source_file": c.get("source_file", ""), "source_function": c.get("source_function", "")}
            registered[pid] = _contract_with_registry_defaults(pid, runtime_row)
    contracts = []
    control_breakdown = []
    run_results = []
    controls_summary = []
    discovery = []
    for pid in sorted(registered):
        contract = _contract_with_registry_defaults(pid, registered[pid])
        calls = by_prompt.get(pid, [])
        call_count = len(calls)
        success_count = sum(1 for c in calls if _to_bool(c.get("prompt_success")))
        failed_count = call_count - success_count
        expected = sum(_safe_int(c.get("expected_items"), 0) for c in calls)
        received = sum(_safe_int(c.get("received_items"), 0) for c in calls)
        json_count = sum(1 for c in calls if _to_bool(c.get("json_valid")))
        schema_count = sum(1 for c in calls if _to_bool(c.get("schema_valid")))
        trace_applicable = [c for c in calls if c.get("traceability_ok") not in (None, "", "not_applicable")]
        trace_ok = sum(1 for c in trace_applicable if _to_bool(c.get("traceability_ok")))
        retry_count = sum(_safe_int(c.get("retry_count"), 0) for c in calls)
        fallback_count = sum(1 for c in calls if _to_bool(c.get("fallback_used")))
        hashes = sorted({str(c.get("prompt_hash") or "") for c in calls if c.get("prompt_hash")})
        observed_schema_hashes = sorted({str(c.get("schema_hash") or "") for c in calls if c.get("schema_hash")})
        contracts.append({
            "prompt_id": pid,
            "prompt_name": contract.get("prompt_name", ""),
            "prompt_scope": contract.get("prompt_scope", ""),
            "prompt_category": contract.get("prompt_category", ""),
            "source_file": contract.get("source_file", ""),
            "source_function": contract.get("source_function", ""),
            "registration_status": contract.get("registration_status", "registered"),
            "executed_in_this_run": "true" if call_count else "false",
            "call_count": call_count,
            "role_in_workflow": contract.get("role_in_workflow") or contract.get("role_description") or "Registered prompt contract.",
            "system_prompt_transcript": contract.get("system_prompt_transcript", ""),
            "user_payload_contract": contract.get("user_payload_contract", ""),
            "required_output_schema": contract.get("required_output_schema", ""),
            "control_clauses": contract.get("control_clauses", ""),
            "validation_mechanism": contract.get("validation_mechanism", ""),
            "prompt_contract_hash": contract.get("prompt_contract_hash") or _stable_hash({"system_prompt": contract.get("system_prompt_transcript", ""), "schema": contract.get("required_output_schema", ""), "controls": contract.get("control_clauses", "")}),
            "observed_prompt_hashes": ", ".join(hashes),
            "observed_schema_hashes": ", ".join(observed_schema_hashes),
        })
        control_text = str(contract.get("control_clauses") or "")
        control_lines = [line.strip("- ").strip() for line in control_text.splitlines() if line.strip()]
        if not control_lines:
            control_lines = ["No registered control clauses available; see runtime telemetry."]
        for idx, clause in enumerate(control_lines, start=1):
            control_name, metric = _infer_control_name_and_metric(clause)
            control_breakdown.append({
                "prompt_id": pid,
                "prompt_name": contract.get("prompt_name", ""),
                "control_index": idx,
                "control_name": control_name,
                "control_clause": clause,
                "validation_mechanism": contract.get("validation_mechanism", ""),
                "metric_used": metric,
                "result_in_this_run": "not_executed" if not call_count else ("observed" if failed_count == 0 else "observed_with_rejections_or_fallback"),
                "plain_language_interpretation": "This clause constrains what the LLM may return and is checked through the listed validation mechanism.",
            })
        controls_summary.append({
            "prompt_id": pid,
            "prompt_name": contract.get("prompt_name", ""),
            "executed_in_this_run": "true" if call_count else "false",
            "json_required": "true" if "json" in (contract.get("control_clauses", "") + contract.get("system_prompt_transcript", "")).lower() else "false",
            "schema_required": "true" if contract.get("required_output_schema") else "false",
            "identifier_preservation_required": "true" if any(x in (contract.get("control_clauses", "") + contract.get("user_payload_contract", "")).lower() for x in ["puid", "item_id", "id"] ) else "false",
            "grounding_required": "true" if any(x in (contract.get("control_clauses", "") + contract.get("system_prompt_transcript", "")).lower() for x in ["use only", "supplied", "provided", "evidence"] ) else "false",
            "no_invent_evidence_required": "true" if "do not invent" in (contract.get("control_clauses", "") + contract.get("system_prompt_transcript", "")).lower() else "false",
            "no_result_change_required": "true" if "precomputed result" in (contract.get("control_clauses", "") + contract.get("system_prompt_transcript", "")).lower() else "false",
            "validation_mechanism": contract.get("validation_mechanism", ""),
        })
        run_results.append({
            "prompt_id": pid,
            "prompt_name": contract.get("prompt_name", ""),
            "prompt_scope": contract.get("prompt_scope", ""),
            "prompt_category": contract.get("prompt_category", ""),
            "executed_in_this_run": "true" if call_count else "false",
            "call_count": call_count,
            "successful_call_count": success_count,
            "failed_call_count": failed_count,
            "prompt_success_rate": _rate(success_count, call_count),
            "json_valid_rate": _rate(json_count, call_count),
            "schema_valid_rate": _rate(schema_count, call_count),
            "traceability_ok_rate": _rate(trace_ok, len(trace_applicable)),
            "completion_rate": _rate(received, expected),
            "retry_count": retry_count,
            "fallback_call_count": fallback_count,
            "expected_items": expected,
            "received_items": received,
            "run_conclusion": "not executed in this run" if call_count == 0 else ("accepted without rejected prompt calls" if failed_count == 0 else "one or more prompt calls required retry, fallback, repair, or were rejected"),
        })
        if contract.get("registration_status") == "unregistered":
            discovery.append({"event_type": "unregistered_prompt", "severity": "warning", "prompt_id": pid, "source_file": contract.get("source_file", ""), "source_function": contract.get("source_function", ""), "message": "Runtime prompt was not in registered inventory."})
    inventory_hash = _stable_hash([{k: r.get(k, "") for k in ["prompt_id", "prompt_contract_hash", "observed_prompt_hashes", "call_count", "registration_status"]} for r in contracts])
    return contracts, control_breakdown, run_results, controls_summary, discovery, inventory_hash


def _build_llm_validation_summary(metrics: Dict[str, Any], prompt_run_results: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    prompt_calls = sum(_safe_int(r.get("call_count"), 0) for r in prompt_run_results)
    successful_prompt_calls = sum(_safe_int(r.get("successful_call_count"), 0) for r in prompt_run_results)
    failed_prompt_calls = sum(_safe_int(r.get("failed_call_count"), 0) for r in prompt_run_results)
    rows = [
        ("json_valid_rate", metrics["json_valid_count"], metrics["num_llm_calls"], metrics["json_valid_rate"], f"Rate of LLM calls that returned parseable JSON. Numeric rule: {NUMERIC_DECISION_RULES['json_valid_rate']}"),
        ("schema_valid_rate", metrics["schema_valid_count"], metrics["num_llm_calls"], metrics["schema_valid_rate"], f"Rate of LLM calls with the required output structure. Numeric rule: {NUMERIC_DECISION_RULES['schema_valid_rate']}"),
        ("completion_rate", metrics["received_items"], metrics["expected_items"], metrics["completion_rate"], f"Accepted items divided by expected items. Numeric rule: {NUMERIC_DECISION_RULES['completion_rate']}"),
        ("traceability_ok_rate", metrics["traceability_ok"], metrics["traceability_expected"], metrics["traceability_ok_rate"], f"Items that preserved expected identifiers where traceability applies. Numeric rule: {NUMERIC_DECISION_RULES['traceability_ok_rate']}"),
        ("fallback_rate", metrics["fallback_items"], metrics["num_requirements"], metrics["fallback_rate"], f"Deterministic fallback use for requirement-level justifications. Numeric rule: {NUMERIC_DECISION_RULES['fallback_rate']}"),
        ("retry_rate", metrics["retry_count"], metrics["num_llm_calls"], metrics["retry_rate"], f"Retries used for LLM call recovery. Numeric rule: {NUMERIC_DECISION_RULES['retry_rate']}"),
        ("prompt_success_rate", successful_prompt_calls, prompt_calls, _rate(successful_prompt_calls, prompt_calls), f"Prompt calls accepted after all applicable validations. Numeric rule: {NUMERIC_DECISION_RULES['prompt_success_rate']}"),
        ("failed_prompt_call_count", failed_prompt_calls, prompt_calls, failed_prompt_calls, "Numeric count rule: target = 0; review-required = 1; critical >= 2."),
    ]
    return [{"metric": m, "numerator": n, "denominator": d, "value": v, "plain_language_interpretation": interp} for m, n, d, v, interp in rows]



def _formula_value(value: Any) -> str:
    if value in (None, ""):
        return "not_available"
    if isinstance(value, float):
        return f"{value:.6f}".rstrip("0").rstrip(".") if not value.is_integer() else str(int(value))
    return str(value)


def _division_formula(metric: str, numerator_label: str, denominator_label: str, row: Dict[str, Any]) -> str:
    n = _formula_value(row.get("numerator"))
    d = _formula_value(row.get("denominator"))
    v = _formula_value(row.get("value"))
    return f"{metric} = {numerator_label} / {denominator_label} = {n} / {d} = {v}"


def _current_run_formula_rows(llm_summary: List[Dict[str, Any]]) -> List[List[Any]]:
    labels = {
        "json_valid_rate": ("json_valid_count", "num_llm_calls"),
        "schema_valid_rate": ("schema_valid_count", "num_llm_calls"),
        "completion_rate": ("received_items", "expected_items"),
        "traceability_ok_rate": ("traceability_ok_items", "traceability_expected_items"),
        "fallback_rate": ("fallback_items", "num_requirements"),
        "retry_rate": ("retry_count", "num_llm_calls"),
        "prompt_success_rate": ("successful_prompt_calls", "executed_prompt_calls"),
    }
    rows: List[List[Any]] = []
    for r in llm_summary:
        metric = str(r.get("metric") or "")
        if metric == "failed_prompt_call_count":
            rows.append([
                metric,
                f"failed_prompt_call_count = failed_prompt_calls = {_formula_value(r.get('numerator'))}",
                "Counts prompt calls rejected by validation or requiring deterministic fallback.",
            ])
            continue
        if metric in labels:
            num_label, den_label = labels[metric]
            rows.append([metric, _division_formula(metric, num_label, den_label, r), r.get("plain_language_interpretation", "")])
    return rows


def _prompt_level_formula_rows() -> List[List[str]]:
    return [
        ["prompt_success_rate", "prompt_success_rate = successful_call_count / call_count", "successful_call_count counts prompt calls that passed JSON, schema, traceability when applicable, and fallback checks."],
        ["json_valid_rate", "json_valid_rate = json_valid_call_count / call_count", "json_valid_call_count counts prompt calls whose output was parseable as JSON."],
        ["schema_valid_rate", "schema_valid_rate = schema_valid_call_count / call_count", "schema_valid_call_count counts prompt calls satisfying the expected output fields."],
        ["traceability_ok_rate", "traceability_ok_rate = traceability_ok_call_count / traceability_applicable_call_count", "Traceability applies when the prompt must preserve PUID, item_id, finding_id, or another supplied identifier."],
        ["completion_rate", "completion_rate = received_items / expected_items", "Used when a prompt returns multiple items, for example one item per PUID or treatment item."],
        ["run_conclusion", "run_conclusion = accepted if failed_call_count = 0; otherwise review/retry/fallback noted", "A prompt can be registered and not executed in a run; in that case rates are not_applicable."],
    ]


def _prompt_scope_definition_rows() -> List[List[str]]:
    return [
        ["audit_matrix", "Requirement-level audit matrix stage. It drafts English justifications for precomputed PUID results and does not adjudicate yes/no/n/a."],
        ["audit_summary", "Audit Summary report-generation stage. It drafts executive, technical, pattern, and treatment-plan narrative sections from supplied evidence."],
        ["audit_summary_repair", "Auxiliary repair stage. It runs only when validation detects missing structured fields in a treatment-plan response."],
    ]


def _stability_formula_rows() -> List[List[str]]:
    return [
        ["comparable_run_count", "comparable_run_count = count(runs where repository, commit_sha, input_fingerprint, prompt_inventory_hash, llm_config_hash, num_requirements and PUID set match)", "Number of repeated executions accepted for stability calculations. Numeric rule: minimum n >= 2; recommended n >= 5 when the protocol requires five executions."],
        ["exact_matrix_agreement_rate", "exact_matrix_agreement_rate = mode_count(compliance_matrix_hash) / comparable_run_count", "Share of comparable runs with the most frequent complete result matrix. Numeric rule: target = 1.000; review-required >= 0.950 and < 1.000; critical < 0.950."],
        ["requirement_result_agreement_rate", "requirement_result_agreement_rate = pairwise_matching_results / pairwise_total_comparisons", "pairwise_matching_results = count of run pairs where result(run_i,puid) = result(run_j,puid), summed across all PUIDs. Numeric rule: acceptable >= 0.990; review-required >= 0.950 and < 0.990; critical < 0.950."],
        ["changed_requirement_count", "changed_requirement_count = count(PUID where count(unique_results_across_comparable_runs) > 1)", "Number of requirements whose yes/no/n/a result changed across comparable runs. Numeric rule: target = 0; review-required 1 to 2; critical >= 3."],
        ["yes_count_mean", "yes_count_mean = sum(yes_count_i) / n", "Mean number of yes results across n comparable runs. Numeric rule: descriptive metric; evaluate together with yes_count_sd, no_count_sd and na_count_sd."],
        ["yes_count_sd", "yes_count_sd = sqrt(sum((yes_count_i - yes_count_mean)^2) / (n - 1))", "Sample standard deviation of yes counts across comparable runs. Same formula applies to no_count_sd and na_count_sd. Numeric rule: target = 0.000; review-required > 0.000 and <= 1.000; critical > 1.000."],
        ["prompt_inventory_stability", "prompt_inventory_stability = mode_count(prompt_inventory_hash) / comparable_run_count", "Confirms whether all comparable runs used the same centralized prompt registry and observed prompt hashes. Numeric rule: target = 1.000; review-required >= 0.950 and < 1.000; critical < 0.950."],
        ["llm_config_stability", "llm_config_stability = mode_count(llm_config_hash) / comparable_run_count", "Confirms whether all comparable runs used the same LLM runtime configuration. Numeric rule: target = 1.000; review-required >= 0.950 and < 1.000; critical < 0.950."],
        ["validation_rate_mean", "validation_rate_mean = sum(validation_rate_i) / n", "Mean of a per-run validation rate such as json_valid_rate, schema_valid_rate, completion_rate, traceability_ok_rate, fallback_rate, retry_rate, or prompt_success_rate. Numeric rule: positive rates acceptable >= 0.950; fallback mean target <= 0.050; retry mean target <= 0.200."],
        ["validation_rate_sd", "validation_rate_sd = sqrt(sum((validation_rate_i - validation_rate_mean)^2) / (n - 1))", "Sample standard deviation of the selected validation rate across comparable runs. Numeric rule: target <= 0.020; review-required > 0.020 and <= 0.050; critical > 0.050."],
        ["Fleiss kappa", "kappa = (P_bar - P_e) / (1 - P_e)", "P_bar is observed categorical agreement across PUID results; P_e is expected agreement from category proportions. Numeric rule: strong agreement >= 0.800; moderate >= 0.600 and < 0.800; weak < 0.600; not_applicable when unsuitable."],
    ]


def _package_contents_rows(telemetry_components: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    included = {c["component"] for c in PACKAGE_COMPONENTS}
    rows = []
    for component in PACKAGE_COMPONENTS:
        rows.append({**component, "included": "true"})
    for t in telemetry_components:
        if t["component"] not in included:
            rows.append({"component": t["component"], "type": t.get("role", "raw_telemetry"), "purpose": "Additional prompt telemetry supplied to the package builder.", "reader_priority": "appendix_raw", "duplication_policy": "Summarized into prompt sheets.", "included": "true"})
    return rows


def _input_evidence_rows(wb: Any, args: argparse.Namespace, telemetry_components: List[Dict[str, Any]], output_xlsx: Path, output_pdf: Path, manifest_path: Path) -> List[Dict[str, Any]]:
    rows = []
    def add(name: str, group: str, role: str, path_value: str, purpose: str, stability_key: str = "yes") -> None:
        p = Path(path_value).expanduser() if path_value else Path("")
        typ, sha, count = _artifact_hash(p) if path_value else ("missing", "", 0)
        availability = "available" if sha else "missing"
        if name == "run_metrics_methodology_pdf" and not sha:
            typ = "file"
            availability = "generated_by_this_step"
            purpose = purpose + " Final SHA-256 is recorded in run-metrics-manifest.json after PDF rendering."
        rows.append({"input_name": name, "evidence_group": group, "evidence_role": role, "artifact_type": typ, "path": str(p) if path_value else "", "file_count": count, "sha256": sha, "availability": availability, "purpose": purpose, "used_for_stability_comparability": stability_key})
    for raw in _read_rows(wb, "input_hashes"):
        rows.append({
            "input_name": raw.get("input_name") or raw.get("name") or raw.get("artifact") or "source_workbook_input",
            "evidence_group": raw.get("component_group") or raw.get("group") or "ai_requirements_excel",
            "evidence_role": "audit_input",
            "artifact_type": raw.get("artifact_type") or "recorded_by_source_workbook",
            "path": raw.get("path") or "",
            "file_count": raw.get("file_count") or "",
            "sha256": raw.get("sha256") or "",
            "availability": raw.get("availability") or "recorded_by_source_workbook",
            "purpose": raw.get("purpose") or "Input hash recorded by the source run workbook.",
            "used_for_stability_comparability": "yes",
        })
    add("technical_inputs_root", "technical_evidence", "technical_input", args.technical_root, "Root folder containing technical artifacts used by Audit Summary.")
    add("vision360_bundle", "technical_evidence", "technical_input", args.vision360_bundle, "Vision360 normalized fingerprint bundle or artifact.")
    add("trivy_payload", "technical_evidence", "technical_input", args.trivy_payload, "Trivy dependency and vulnerability evidence.")
    add("mobsf_report", "technical_evidence", "technical_input", args.mobsf_report, "MobSF static analysis evidence.")
    add("mobsf_dynamic_report", "technical_evidence", "technical_input", args.mobsf_dynamic_report, "MobSF dynamic analysis evidence.")
    add("sast_findings", "technical_evidence", "technical_input", args.sast_findings, "SAST findings evidence.")
    add("audit_summary_analysis_pack", "audit_summary", "generated_audit_input", args.analysis_pack, "Stage 1 analysis pack consumed by the Audit Summary generator.")
    add("audit_summary_literals", "audit_summary", "configuration_input", args.audit_config, "Application metadata and report literal configuration.")
    add("audit_summary_docx", "audit_summary", "generated_output", args.audit_summary_docx, "Generated Audit Summary DOCX.", "no")
    add("audit_summary_pdf", "audit_summary", "generated_output", args.audit_summary_pdf, "Generated Audit Summary PDF.", "no")
    for t in telemetry_components:
        p = Path(t.get("path", ""))
        rows.append({"input_name": p.stem, "evidence_group": "prompt_telemetry", "evidence_role": "raw_prompt_telemetry", "artifact_type": "file", "path": str(p), "file_count": 1, "sha256": t.get("sha256", ""), "availability": "included", "purpose": "Raw prompt-call telemetry used to reconstruct prompt execution and validation evidence.", "used_for_stability_comparability": "yes"})
    add("run_metrics_workbook", "run_metrics", "package_component", str(output_xlsx), "Generated run-metrics workbook.", "no")
    add("run_metrics_methodology_pdf", "run_metrics", "package_component", str(output_pdf), "Generated methodology PDF.", "no")
    add("run_metrics_manifest", "run_metrics", "package_component", str(manifest_path), "Generated package manifest.", "no")
    contracts_path = output_xlsx.parent / PROMPT_CONTRACTS_PACKAGE_PATH
    add("prompt_contracts_registry", "run_metrics", "prompt_contract_registry", str(contracts_path), "Central JSON registry with prompt transcripts, schemas, controls and validation mechanisms.", "yes")
    validator_path = output_xlsx.parent / PROMPT_REGISTRY_VALIDATOR_PACKAGE_PATH
    add("prompt_registry_validator", "run_metrics", "package_tool", str(validator_path), "Tool that validates prompt registry completeness and LLM-call coverage.", "no")
    tool_path = output_xlsx.parent / STABILITY_TOOL_PACKAGE_PATH
    add("build_stability_analysis_tool", "run_metrics", "package_tool", str(tool_path), "Tool that generates stability-analysis.xlsx from n run-metrics packages.", "no")
    # de-duplicate by input_name/path/sha
    seen = set(); out = []
    for r in rows:
        key = (str(r.get("input_name")), str(r.get("path")), str(r.get("sha256")))
        if key not in seen:
            seen.add(key); out.append(r)
    return out

# ---------------------------------------------------------------------------
# Workbook writing
# ---------------------------------------------------------------------------

def _replace_sheet(wb: Any, name: str) -> Any:
    safe = name[:31]
    if safe in wb.sheetnames:
        del wb[safe]
    return wb.create_sheet(safe)


def _write_rows(wb: Any, name: str, rows: List[Dict[str, Any]], headers: List[str]) -> None:
    ws = _replace_sheet(wb, name)
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])


def _write_kv(wb: Any, name: str, rows: List[Tuple[str, Any]]) -> None:
    ws = _replace_sheet(wb, name)
    ws.append(["key", "value"])
    for k, v in rows:
        ws.append([k, v])


def _style_workbook(wb: Any) -> None:
    order = [
        "user_guide", "metric_definitions", "run_summary", "adjudication_summary", "adjudication_by_requirement", "adjudication_by_flag", "package_contents", "input_evidence", "prompt_contracts", "prompt_control_breakdown", "prompt_run_results", "llm_validation_summary", "stability_readiness", "Appendix_Requirement_Rows", "Appendix_LLM_Calls", "Appendix_LLM_Items", "Appendix_Prompt_Calls", "Appendix_Discovery", "Appendix_Package_Manifest"
    ]
    for idx, name in enumerate(reversed(order)):
        if name in wb.sheetnames:
            ws = wb[name]
            wb._sheets.remove(ws)
            wb._sheets.insert(0, ws)
    header_fill = PatternFill("solid", fgColor=THEME_NAVY)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    body_font = Font(name="Arial", color=THEME_TEXT, size=9)
    side = Side(style="thin", color="D0D7DE")
    border = Border(left=side, right=side, top=side, bottom=side)
    existing_tables = set()
    for ws in wb.worksheets:
        ws.sheet_view.showGridLines = False
        ws.freeze_panes = "A2"
        max_row, max_col = ws.max_row, ws.max_column
        if max_row < 1 or max_col < 1:
            continue
        ws.row_dimensions[1].height = 26
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for r in range(2, max_row + 1):
            for c in range(1, max_col + 1):
                cell = ws.cell(r, c)
                cell.font = body_font
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = border
                if r % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor=THEME_LIGHT)
                if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool):
                    cell.number_format = "#,##0.000" if isinstance(cell.value, float) and not math.isclose(cell.value, round(cell.value)) else "#,##0"
        for c in range(1, max_col + 1):
            header = str(ws.cell(1, c).value or "").lower()
            width = 18
            if "transcript" in header or "schema" in header or "contract" in header or "clause" in header or "mechanism" in header or "interpretation" in header or "purpose" in header or "guide" in header:
                width = 50
            elif "hash" in header or "sha" in header:
                width = 36
            elif "path" in header:
                width = 45
            elif "prompt" in header:
                width = 26
            ws.column_dimensions[get_column_letter(c)].width = width
        if max_row >= 2 and max_col >= 2:
            try:
                base = "tbl_" + "".join(ch if ch.isalnum() else "_" for ch in ws.title.lower())[:28]
                name = base; i = 2
                while name in existing_tables:
                    name = f"{base}_{i}"; i += 1
                existing_tables.add(name)
                tab = Table(displayName=name, ref=f"A1:{get_column_letter(max_col)}{max_row}")
                tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
                ws.add_table(tab)
            except Exception:
                pass
        if ws.title.startswith("Appendix_"):
            ws.sheet_properties.tabColor = "A5A5A5"
        else:
            ws.sheet_properties.tabColor = THEME_NAVY


def _manifest_rows(out_dir: Path, raw_path: Path) -> List[Dict[str, Any]]:
    rows = []
    for rel in [RUN_METRICS_XLSX, RUN_METRICS_PDF, RUN_METRICS_MANIFEST, PROMPT_CONTRACTS_PACKAGE_PATH, PROMPT_REGISTRY_VALIDATOR_PACKAGE_PATH, STABILITY_TOOL_PACKAGE_PATH, "telemetry/prompt-telemetry-ai-requirements.jsonl", "telemetry/prompt-telemetry-audit-summary.jsonl"]:
        p = out_dir / rel
        if p.is_file():
            rows.append({"artifact": rel, "path": str(p), "sha256": _sha256_file(p), "size_bytes": p.stat().st_size})
    rows.append({"artifact": "raw_metrics_source", "path": str(raw_path), "sha256": _sha256_file(raw_path), "size_bytes": raw_path.stat().st_size if raw_path.is_file() else 0})
    rows.append({"artifact": "package_generated_at_utc", "path": _now_utc(), "sha256": "", "size_bytes": ""})
    return rows


def _write_manifest_json(path: Path, out_dir: Path, summary: Dict[str, Any], files: List[Dict[str, Any]]) -> None:
    payload = {
        "manifest_version": "2.0",
        "package_name": "run-metrics",
        "generated_at_utc": _now_utc(),
        "repository": summary.get("repository", ""),
        "commit_sha": summary.get("commit_sha", ""),
        "run_id": summary.get("run_id", ""),
        "workflow": summary.get("workflow", ""),
        "files": [],
    }
    for row in files:
        component = str(row.get("component") or row.get("artifact") or "")
        if not component:
            continue
        p = out_dir / component
        if p.is_file():
            payload["files"].append({"path": component, "role": row.get("type") or row.get("role") or "package_file", "sha256": _sha256_file(p), "size_bytes": p.stat().st_size})
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def _build_user_guide() -> List[Tuple[str, Any]]:
    return [
        ("what_this_workbook_is", "This workbook documents one audit execution. It explains the run, the inputs, the LLM prompt contracts, how outputs were controlled and validated, and which data can later be aggregated across repeated executions."),
        ("recommended_reading_order", "Read user_guide first, then metric_definitions, run_summary, adjudication_summary, adjudication_by_requirement, adjudication_by_flag, input_evidence, prompt_contracts, prompt_control_breakdown, prompt_run_results, llm_validation_summary, and stability_readiness. Appendix sheets preserve raw evidence."),
        ("appendix_policy", "Appendix sheets are technical backup. They are not the starting point for interpretation, but they allow reconstruction of requirement rows, LLM calls, item-level validation, prompt calls, discovery warnings and file hashes."),
        ("what_the_llm_can_do", "The LLM can generate controlled text such as justifications, narratives, pattern writeups and treatment-plan wording from supplied evidence."),
        ("what_the_llm_cannot_do", "The LLM does not compute or change yes/no/n/a compliance results, does not decide flags and must not invent evidence, scanner findings, files, CVEs, packages or versions."),
        ("how_to_check_prompt_control", "Use prompt_contracts to read the actual prompt contract and output schema. Use prompt_control_breakdown to see each control clause, validation mechanism and metric. Use prompt_run_results to see what happened in this execution."),
        ("how_to_build_stability_analysis", "Extract run-metrics.zip from repeated executions into runs/run_01, runs/run_02, ..., runs/run_n. Then run: python tools/build_stability_analysis.py --input-dir runs --output stability-analysis.xlsx."),
    ]


def _build_stability_readiness(summary: Dict[str, Any], input_evidence: List[Dict[str, Any]], prompt_inventory_hash: str, metrics: Dict[str, Any]) -> List[Dict[str, Any]]:
    checks = [
        ("run_id_available", bool(summary.get("run_id")), "A run identifier is needed to distinguish repeated executions."),
        ("repository_available", bool(summary.get("repository")), "Repository identifies the codebase being audited."),
        ("commit_sha_available", bool(summary.get("commit_sha")), "Commit SHA is required for comparability."),
        ("input_evidence_hashes_available", any(r.get("sha256") for r in input_evidence if r.get("used_for_stability_comparability") == "yes"), "Critical input hashes are required to compare executions."),
        ("llm_config_hash_available", bool(summary.get("llm_config_hash")), "LLM configuration hash is required to verify comparable LLM settings."),
        ("prompt_inventory_hash_available", bool(prompt_inventory_hash), "Prompt inventory hash is required to verify comparable prompt contracts."),
        ("compliance_matrix_hash_available", bool(metrics.get("compliance_matrix_hash")), "Compliance matrix hash is required for exact matrix agreement."),
        ("stability_tool_packaged", (STABILITY_TOOL_SRC.is_file()), "tools/build_stability_analysis.py must be included in run-metrics.zip."),
    ]
    return [{"check": name, "status": "ready" if ok else "missing", "meaning": text} for name, ok, text in checks]

# ---------------------------------------------------------------------------
# PDF generation
# ---------------------------------------------------------------------------

class PdfBuilder:
    def __init__(self, path: Path):
        self.path = path
        self.styles = getSampleStyleSheet()
        self.story: List[Any] = []
        self.table_no = 0
        self.figure_no = 0
        self.styles.add(ParagraphStyle("Small", parent=self.styles["BodyText"], fontName="Helvetica", fontSize=7.3, leading=9, textColor=colors.HexColor("#1F2933")))
        # ReportLab TableStyle TEXTCOLOR does not reliably override the text color
        # of Paragraph objects already placed in table cells. Table headers are
        # therefore rendered with their own white, bold Paragraph style. This
        # keeps header text legible on the dark-blue header background in the
        # generated PDF.
        self.styles.add(ParagraphStyle("TableHeaderSmall", parent=self.styles["BodyText"], fontName="Helvetica-Bold", fontSize=7.3, leading=9, textColor=colors.white))
        self.styles.add(ParagraphStyle("Caption", parent=self.styles["BodyText"], fontName="Helvetica-Oblique", fontSize=8, leading=10, textColor=colors.HexColor("#4B5563")))
        self.styles.add(ParagraphStyle("TitleCenter", parent=self.styles["Title"], alignment=TA_CENTER, textColor=colors.HexColor("#17365D")))
        self.styles.add(ParagraphStyle("H1", parent=self.styles["Heading1"], textColor=colors.HexColor("#17365D")))
        self.styles.add(ParagraphStyle("H2", parent=self.styles["Heading2"], textColor=colors.HexColor("#17365D")))

    def para(self, text: Any, style: str = "BodyText") -> None:
        self.story.append(Paragraph(str(text or "").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"), self.styles[style]))
        self.story.append(Spacer(1, 0.06 * inch))

    def h1(self, text: str) -> None:
        self.story.append(Paragraph(text, self.styles["H1"])); self.story.append(Spacer(1, 0.08 * inch))

    def table(self, title: str, headers: List[str], rows: List[List[Any]], widths: List[float] | None = None, max_rows: int = 18) -> None:
        self.table_no += 1
        self.story.append(Paragraph(f"Table {self.table_no}. {title}", self.styles["Caption"]))
        rows = rows[:max_rows]
        data = [[Paragraph(str(h), self.styles["TableHeaderSmall"]) for h in headers]]
        for row in rows:
            data.append([Paragraph(str(c or ""), self.styles["Small"]) for c in row])
        page_width = A4[0] - 0.8 * inch
        col_widths = [page_width * w for w in widths] if widths else None
        table = RLTable(data, colWidths=col_widths, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#17365D")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("GRID", (0,0), (-1,-1), 0.35, colors.HexColor("#D0D7DE")),
            ("VALIGN", (0,0), (-1,-1), "TOP"),
            ("FONTSIZE", (0,0), (-1,-1), 7),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F7F9FC")]),
        ]))
        self.story.append(table)
        self.story.append(Spacer(1, 0.12 * inch))

    def fig(self, img_path: Path, title: str, caption: str) -> None:
        self.figure_no += 1
        self.story.append(Paragraph(f"Figure {self.figure_no}. {title}", self.styles["Caption"]))
        self.story.append(Image(str(img_path), width=6.7 * inch, height=3.4 * inch))
        self.para(caption, "Caption")

    def build(self) -> None:
        doc = SimpleDocTemplate(str(self.path), pagesize=A4, rightMargin=0.4*inch, leftMargin=0.4*inch, topMargin=0.45*inch, bottomMargin=0.45*inch)
        doc.build(self.story)


def _make_bar_chart(path: Path, labels: List[str], values: List[float], title: str, ylabel: str, percent: bool = False) -> None:
    fig, ax = plt.subplots(figsize=(7.8, 3.9))
    bars = ax.bar(labels, values)
    ax.set_title(title, fontsize=12, fontweight="bold")
    ax.set_ylabel(ylabel)
    if percent:
        ax.set_ylim(0, 1.05)
        ax.yaxis.set_major_formatter(lambda x, pos: f"{x:.0%}")
        for b, v in zip(bars, values):
            ax.text(b.get_x()+b.get_width()/2, min(1.02, v+0.025), f"{v:.1%}", ha="center", fontsize=8)
    else:
        ymax = max(values + [1])
        for b, v in zip(bars, values):
            ax.text(b.get_x()+b.get_width()/2, v + ymax*0.02, str(int(v)), ha="center", fontsize=8)
    ax.grid(axis="y", linestyle="--", linewidth=0.4, alpha=0.35)
    fig.tight_layout(pad=1.2)
    fig.savefig(path, dpi=180, bbox_inches="tight")
    plt.close(fig)


def _generate_pdf(path: Path, metrics: Dict[str, Any], rows: Dict[str, List[Dict[str, Any]]], chart_dir: Path) -> None:
    chart_dir.mkdir(parents=True, exist_ok=True)
    b = PdfBuilder(path)
    summary = metrics.get("summary", {})
    b.story.append(Paragraph("Run Metrics Methodology", b.styles["TitleCenter"]))
    b.story.append(Paragraph("User-oriented evidence package for LLM-controlled audit reproducibility", b.styles["TitleCenter"]))
    b.table("Execution identity", ["Field", "Value"], [["Run ID", summary.get("run_id", "")], ["Repository", summary.get("repository", "")], ["Commit SHA", summary.get("commit_sha", "")], ["Workflow", summary.get("workflow", "")], ["Generated at", summary.get("generated_at_utc", summary.get("package_generated_at_utc", ""))], ["LLM configuration hash", summary.get("llm_config_hash", "")], ["Compliance matrix hash", metrics.get("compliance_matrix_hash", "")]], widths=[0.30,0.70])
    b.para("This PDF explains every component in run-metrics.zip, the workbook reading order, the metrics used, the role of each LLM prompt, the output-control and validation mechanisms observed in this run, and how the included tool generates stability-analysis.xlsx from n repeated runs.")
    b.h1("1. How to read this package")
    b.para("The package documents one audit execution. It is not an alternative audit engine. The compliance results yes, no and n/a are computed by the audit workflow and recorded here. LLM metrics document controlled text generation, JSON/schema validation, traceability checks, retries, repairs and fallback. They do not grant the LLM authority to change deterministic compliance outcomes.")
    b.table("Run Metrics package contents", ["Component", "Type", "Purpose"], [[r["component"], r["type"], r["purpose"]] for r in rows["package_contents"]], widths=[0.28,0.18,0.54], max_rows=10)
    b.table("Visible workbook reading order", ["Sheet", "What the user should learn"], [["user_guide", "How to read the workbook and which sheets are primary or Appendix."], ["metric_definitions", "Plain-language definition, formula and interpretation for each metric."], ["run_summary", "Current execution identity, counts and hashes."], ["adjudication_summary", "Contradiction, risk-precedence and partial-evidence counters."], ["adjudication_by_requirement", "Requirement-level evidence adjudication diagnostics."], ["adjudication_by_flag", "Flag-level support, contradiction, unknown and missing evidence roles."], ["input_evidence", "Inputs, generated outputs, telemetry and package components with hashes."], ["prompt_contracts", "Prompt IDs, exact prompt transcripts, required schemas and validation mechanisms."], ["prompt_control_breakdown", "Each control clause mapped to a validation mechanism and metric."], ["prompt_run_results", "Prompt-level outcome for this execution."], ["llm_validation_summary", "Aggregated LLM-control metrics."], ["stability_readiness", "Whether this run contains the fields needed for repeated-run comparison."], ["Appendix_*", "Raw evidence preserved for reconstruction and traceability."]], widths=[0.28,0.72], max_rows=12)
    b.h1("2. Metrics used")
    b.table("Metric formulas, thresholds and interpretation", ["Metric", "Question", "Formula", "Numeric decision rule", "Interpretation"], [[r["metric"], r["plain_language_question"], r["formula"], r.get("numeric_decision_rule", _metric_decision_rule(r["metric"])), r["interpretation"]] for r in METRIC_DEFINITIONS], widths=[0.16,0.24,0.18,0.24,0.18], max_rows=20)
    b.table("Numeric decision thresholds for single-run metrics", ["Metric family", "Target or acceptable", "Review-required", "Critical"], [["Positive LLM validation rates: json_valid_rate, schema_valid_rate, completion_rate, traceability_ok_rate, prompt_success_rate", ">= 0.950", ">= 0.900 and < 0.950", "< 0.900"], ["fallback_rate", "<= 0.050", "> 0.050 and <= 0.100", "> 0.100"], ["retry_rate", "<= 0.200", "> 0.200 and <= 0.500", "> 0.500"], ["failed_prompt_call_count", "0", "1", ">= 2"]], widths=[0.40,0.20,0.22,0.18], max_rows=8)

    b.h1("3. Current execution summary")
    compliance_chart = chart_dir / "compliance.png"
    _make_bar_chart(compliance_chart, ["yes", "no", "n/a"], [metrics["counts"].get("yes",0), metrics["counts"].get("no",0), metrics["counts"].get("n/a",0)], "Compliance result distribution", "Requirement count")
    b.fig(compliance_chart, "Compliance result distribution", "This figure summarizes the deterministic results recorded in Appendix_Requirement_Rows. It does not recalculate or reinterpret audit outcomes.")
    validation_chart = chart_dir / "validation.png"
    vals = []
    for key in ["json_valid_rate", "schema_valid_rate", "completion_rate", "traceability_ok_rate"]:
        value = metrics.get(key)
        vals.append(0 if value == "not_applicable" else _safe_float(value))
    _make_bar_chart(validation_chart, ["JSON", "schema", "completion", "traceability"], vals, "LLM output validation rates", "Rate", percent=True)
    b.fig(validation_chart, "LLM output validation rates", "This figure shows whether LLM outputs were machine-readable, structurally valid, complete, and traceable where traceability applies.")

    b.h1("4. Evidence adjudication diagnostics")
    b.para("This section documents how the deterministic compliance layer handled mapped evidence before any narrative LLM stage. These diagnostics do not change the yes/no/n/a matrix. They explain whether a requirement was supported, contradicted by mapped risk evidence, limited by partial evidence, or affected by missing or unknown flag evidence.")
    b.table("Deterministic adjudication policy", ["Evidence situation", "How it is handled", "Where to inspect it"], [
        ["Mapped risk signal contradicts the expected secure posture", "Risk precedence is applied and the requirement cannot be reported as compliant.", "adjudication_by_requirement and adjudication_by_flag"],
        ["All relevant non-applicability flags support the expected posture", "The requirement can be reported as compliant when no contradictory signal is present.", "adjudication_by_requirement"],
        ["Evidence is missing, unknown, or only partially supports the criterion", "The result is treated as non-conclusive or not applicable according to the requirement context; it is not used to assert compliance.", "adjudication_summary and adjudication_by_requirement"],
        ["Multiple equivalent findings support the same flag", "They may increase evidence_count and traceability, but do not multiply the compliance decision.", "adjudication_by_flag and source evidence rows"],
    ], widths=[0.28,0.42,0.30], max_rows=6)
    adj_summary = metrics.get("adjudication_summary", {}) or {}
    if adj_summary:
        adj_meanings = {
            "contradictory_requirement_count": "Requirements with at least one mapped signal contradicting the expected secure posture.",
            "contradictory_flag_reference_count": "PUID-to-flag references whose observed value contradicted the expected posture.",
            "risk_precedence_count": "Requirements where the conservative risk-precedence rule was applied.",
            "partial_evidence_requirement_count": "Requirements with partial, missing, or unknown evidence without a direct contradiction.",
            "unknown_flag_reference_count": "Mapped flag references evaluated as unknown or non-conclusive.",
            "missing_flag_reference_count": "Mapped flag references declared by requirements but absent from the fingerprint.",
            "adjudicated_requirement_count": "Requirements for which adjudication diagnostics were recorded.",
            "adjudicated_flag_reference_count": "Total PUID-to-flag references inspected by the diagnostics layer.",
        }
        preferred_adj_keys = [
            "adjudicated_requirement_count",
            "adjudicated_flag_reference_count",
            "contradictory_requirement_count",
            "contradictory_flag_reference_count",
            "risk_precedence_count",
            "partial_evidence_requirement_count",
            "unknown_flag_reference_count",
            "missing_flag_reference_count",
        ]
        adj_rows = []
        for key in preferred_adj_keys:
            if key in adj_summary:
                adj_rows.append([key, adj_summary.get(key), adj_meanings.get(key, "Adjudication diagnostic metric.")])
        for key, value in adj_summary.items():
            if key not in preferred_adj_keys:
                adj_rows.append([key, value, adj_meanings.get(key, "Additional adjudication diagnostic metric.")])
        b.table("Adjudication counters recorded in this run", ["Metric", "Value", "Meaning"], adj_rows, widths=[0.34,0.14,0.52], max_rows=12)
    else:
        b.para("No adjudication diagnostic sheets were found in the input workbook. This is expected for older run-metrics workbooks generated before the adjudication-diagnostics patch.", "Body")
    b.table("Adjudication traceability sheets", ["Sheet", "Purpose"], [
        ["adjudication_summary", "Global counters for contradictory evidence, risk precedence, partial evidence, unknown flags and missing flags."],
        ["adjudication_by_requirement", "One row per requirement with decision basis, support count, contradiction count, unknown count, missing count and whether risk precedence was applied."],
        ["adjudication_by_flag", "One row per PUID-to-flag relation with flag classification, observed value, expected value and support, contradiction, unknown or missing outcome."],
    ], widths=[0.30,0.70], max_rows=6)

    b.h1("5. Input evidence and hash controls")
    b.table("Input evidence and package hashes", ["Name", "Role", "Availability", "SHA-256"], [[r.get("input_name"), r.get("evidence_role"), r.get("availability"), r.get("sha256")] for r in rows["input_evidence"]], widths=[0.25,0.20,0.15,0.40], max_rows=16)
    b.h1("6. Role of the LLM and prompt contracts")
    b.para("Each registered prompt has a stable prompt_id. Each runtime invocation receives a prompt_call_id. The workbook records the exact prompt transcript or contract, expected output schema, control clauses, prompt hash, schema hash, payload hash and observed validation result. Dynamic user payloads are not fully duplicated in the visible sheets to avoid an unreadable workbook; they are represented by payload_hash and raw JSONL telemetry.")
    b.table("Prompt roles and execution in this run", ["Prompt", "Scope", "Calls", "Role"], [[r.get("prompt_id"), r.get("prompt_scope"), r.get("call_count"), r.get("role_in_workflow")] for r in rows["prompt_contracts"]], widths=[0.14,0.18,0.10,0.58], max_rows=12)
    b.table("Prompt output-control summary", ["Prompt", "JSON", "Schema", "Identifier", "Grounding", "No invention", "No result change"], [[r.get("prompt_id"), r.get("json_required"), r.get("schema_required"), r.get("identifier_preservation_required"), r.get("grounding_required"), r.get("no_invent_evidence_required"), r.get("no_result_change_required")] for r in rows["prompt_controls"]], widths=[0.15,0.10,0.12,0.15,0.15,0.17,0.16], max_rows=12)
    b.h1("7. Prompt validation results in this execution")
    b.para("This section reports the formulas used for the current run. Rates are calculated directly from the runtime telemetry rows, not estimated from the narrative text.")
    b.table("Prompt-level success summary", ["Prompt", "Executed", "Calls", "Success", "JSON", "Schema", "Traceability", "Conclusion"], [[r.get("prompt_id"), r.get("executed_in_this_run"), r.get("call_count"), r.get("prompt_success_rate"), r.get("json_valid_rate"), r.get("schema_valid_rate"), r.get("traceability_ok_rate"), r.get("run_conclusion")] for r in rows["prompt_run_results"]], widths=[0.13,0.10,0.08,0.10,0.10,0.10,0.13,0.26], max_rows=12)
    b.table("Prompt-level formulas", ["Metric", "Formula", "How to read it"], _prompt_level_formula_rows(), widths=[0.20,0.40,0.40], max_rows=12)
    b.table("Aggregated LLM validation metrics", ["Metric", "Numerator", "Denominator", "Value", "Interpretation"], [[r.get("metric"), r.get("numerator"), r.get("denominator"), r.get("value"), r.get("plain_language_interpretation")] for r in rows["llm_validation_summary"]], widths=[0.20,0.13,0.13,0.12,0.42], max_rows=12)
    b.table("Current-run formulas with applied values", ["Metric", "Formula applied in this run", "Interpretation"], _current_run_formula_rows(rows["llm_validation_summary"]), widths=[0.20,0.47,0.33], max_rows=12)
    scope_chart = chart_dir / "prompt_scope.png"
    scope_counts = Counter(r.get("prompt_scope") for r in rows["prompt_calls"])
    _make_bar_chart(scope_chart, list(scope_counts.keys()) or ["none"], [float(v) for v in (scope_counts.values() or [0])], "Prompt calls by scope", "Prompt calls")
    b.table("Prompt scope definitions", ["Scope", "Meaning"], _prompt_scope_definition_rows(), widths=[0.25,0.75], max_rows=6)
    b.fig(scope_chart, "Prompt calls by scope", "The chart counts executed prompt calls by scope. audit_matrix is the requirement-level justification stage. audit_summary is the report narrative stage. audit_summary_repair is an auxiliary repair scope and appears only when a repair prompt is invoked.")
    b.h1("8. How stability-analysis.xlsx is generated")
    b.para("The package includes tools/build_stability_analysis.py. The user extracts each run-metrics.zip into a folder such as runs/run_01, runs/run_02, ..., runs/run_n. Then the user runs: python tools/build_stability_analysis.py --input-dir runs --output stability-analysis.xlsx. The script reads each run-metrics.xlsx, verifies comparability and produces metrics, tables and Appendix evidence for n repeated executions.")
    b.table("Inputs consumed by build_stability_analysis.py", ["Input", "Required", "Purpose"], [["runs/run_*/run-metrics.xlsx", "yes", "Primary source for run identity, matrix hash, prompt inventory hash, validation rates and requirement rows."], ["run-metrics-manifest.json", "optional", "File integrity reference when present."], ["telemetry/*.jsonl", "optional", "Raw prompt-call reconstruction when a future version chooses to re-derive prompt rows." ]], widths=[0.34,0.13,0.53], max_rows=10)
    b.table("Stability-analysis detailed formulas", ["Metric", "Formula", "Interpretation"], _stability_formula_rows(), widths=[0.22,0.48,0.30], max_rows=20)
    b.h1("9. Interpretation limits")
    b.para("The package records what the workflow observed. Missing technical artifacts are marked as missing rather than inferred. Prompts not executed are documented as registered but not executed. Stability statistics are generated by tools/build_stability_analysis.py from n repeated run-metrics workbooks; they are not invented from a single execution.")
    b.build()

# ---------------------------------------------------------------------------
# Main build
# ---------------------------------------------------------------------------

def build(args: argparse.Namespace) -> None:
    out_dir = Path(args.out_dir).expanduser(); out_dir.mkdir(parents=True, exist_ok=True)
    raw_path = Path(args.raw_metrics).expanduser()
    if not raw_path.is_file():
        raise SystemExit(f"Missing run metrics workbook: {raw_path}")
    output_xlsx = out_dir / RUN_METRICS_XLSX
    output_pdf = out_dir / RUN_METRICS_PDF
    manifest_path = out_dir / RUN_METRICS_MANIFEST
    if raw_path.resolve() != output_xlsx.resolve():
        shutil.copy2(raw_path, output_xlsx)
    # Copy stability tool into the package.
    tool_dst = out_dir / STABILITY_TOOL_PACKAGE_PATH
    tool_dst.parent.mkdir(parents=True, exist_ok=True)
    if STABILITY_TOOL_SRC.is_file():
        shutil.copy2(STABILITY_TOOL_SRC, tool_dst)
    else:
        tool_dst.write_text("# build_stability_analysis.py was not found in the repository at package build time.\n", encoding="utf-8")
    contracts_dst = out_dir / PROMPT_CONTRACTS_PACKAGE_PATH
    contracts_dst.parent.mkdir(parents=True, exist_ok=True)
    if PROMPT_CONTRACTS_FILE_SRC.is_file():
        shutil.copy2(PROMPT_CONTRACTS_FILE_SRC, contracts_dst)
    else:
        contracts_dst.write_text(json.dumps({"registry_version": "missing", "contracts": []}, indent=2), encoding="utf-8")
    validator_dst = out_dir / PROMPT_REGISTRY_VALIDATOR_PACKAGE_PATH
    validator_dst.parent.mkdir(parents=True, exist_ok=True)
    if PROMPT_REGISTRY_VALIDATOR_SRC.is_file():
        shutil.copy2(PROMPT_REGISTRY_VALIDATOR_SRC, validator_dst)
    else:
        validator_dst.write_text("# validate_prompt_registry.py was not found in the repository at package build time.\n", encoding="utf-8")
    wb_raw = openpyxl.load_workbook(output_xlsx, data_only=True)
    metrics = _compute_base_metrics(wb_raw)
    telemetry_events, telemetry_components = _copy_telemetry_files(args.prompt_telemetry or [], out_dir)
    prompt_calls = _prompt_call_rows(telemetry_events, _events_from_existing_workbook(wb_raw))
    prompt_contracts, prompt_control_breakdown, prompt_run_results, prompt_controls, discovery_rows, prompt_inventory_hash = _build_prompt_sections(prompt_calls)
    llm_validation_summary = _build_llm_validation_summary(metrics, prompt_run_results)
    package_contents = _package_contents_rows(telemetry_components)
    # provisional manifest before input evidence finalizes
    _write_manifest_json(manifest_path, out_dir, metrics.get("summary", {}), package_contents)
    input_evidence = _input_evidence_rows(wb_raw, args, telemetry_components, output_xlsx, output_pdf, manifest_path)
    stability_readiness = _build_stability_readiness(metrics.get("summary", {}), input_evidence, prompt_inventory_hash, metrics)
    # Create new workbook from scratch with the reading order requested.
    wb = openpyxl.Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    summary = dict(metrics.get("summary", {}))
    summary.update({
        "package_generated_at_utc": _now_utc(),
        "num_requirements": metrics["num_requirements"],
        "num_yes": metrics["counts"].get("yes", 0),
        "num_no": metrics["counts"].get("no", 0),
        "num_na": metrics["counts"].get("n/a", 0),
        "compliance_matrix_hash": metrics["compliance_matrix_hash"],
        "prompt_inventory_hash": prompt_inventory_hash,
        "prompt_call_count": len(prompt_calls),
        "executed_prompt_contract_count": sum(1 for r in prompt_contracts if r.get("executed_in_this_run") == "true"),
        "successful_prompt_call_count": sum(1 for r in prompt_calls if r.get("prompt_success") == "true"),
    })
    for key, value in metrics.get("adjudication_summary", {}).items():
        summary.setdefault(str(key), value)
    _write_kv(wb, "user_guide", _build_user_guide())
    _write_rows(wb, "metric_definitions", METRIC_DEFINITIONS, ["metric", "plain_language_question", "purpose", "formula", "numerator", "denominator", "source_sheet", "interpretation", "numeric_decision_rule", "single_run", "stability_analysis"])
    _write_kv(wb, "run_summary", list(summary.items()))
    adj_summary_rows = [{"metric": k, "value": v} for k, v in metrics.get("adjudication_summary", {}).items()]
    _write_rows(wb, "adjudication_summary", adj_summary_rows, ["metric", "value"])
    _write_rows(wb, "adjudication_by_requirement", metrics.get("adjudication_by_requirement", []), sorted(set().union(*(r.keys() for r in metrics.get("adjudication_by_requirement", [])))) if metrics.get("adjudication_by_requirement") else ["puid", "result", "decision_basis"])
    _write_rows(wb, "adjudication_by_flag", metrics.get("adjudication_by_flag", []), sorted(set().union(*(r.keys() for r in metrics.get("adjudication_by_flag", [])))) if metrics.get("adjudication_by_flag") else ["puid", "flag_id", "outcome"])
    _write_rows(wb, "package_contents", package_contents, ["component", "type", "included", "reader_priority", "purpose", "duplication_policy"])
    _write_rows(wb, "input_evidence", input_evidence, ["input_name", "evidence_group", "evidence_role", "artifact_type", "path", "file_count", "sha256", "availability", "purpose", "used_for_stability_comparability"])
    _write_rows(wb, "prompt_contracts", prompt_contracts, ["prompt_id", "prompt_name", "prompt_scope", "prompt_category", "source_file", "source_function", "registration_status", "executed_in_this_run", "call_count", "role_in_workflow", "system_prompt_transcript", "user_payload_contract", "required_output_schema", "control_clauses", "validation_mechanism", "prompt_contract_hash", "observed_prompt_hashes", "observed_schema_hashes"])
    _write_rows(wb, "prompt_control_breakdown", prompt_control_breakdown, ["prompt_id", "prompt_name", "control_index", "control_name", "control_clause", "validation_mechanism", "metric_used", "result_in_this_run", "plain_language_interpretation"])
    _write_rows(wb, "prompt_run_results", prompt_run_results, ["prompt_id", "prompt_name", "prompt_scope", "prompt_category", "executed_in_this_run", "call_count", "successful_call_count", "failed_call_count", "prompt_success_rate", "json_valid_rate", "schema_valid_rate", "traceability_ok_rate", "completion_rate", "retry_count", "fallback_call_count", "expected_items", "received_items", "run_conclusion"])
    _write_rows(wb, "llm_validation_summary", llm_validation_summary, ["metric", "numerator", "denominator", "value", "plain_language_interpretation"])
    _write_rows(wb, "stability_readiness", stability_readiness, ["check", "status", "meaning"])
    _write_rows(wb, "Appendix_Requirement_Rows", metrics["compliance"], sorted(set().union(*(r.keys() for r in metrics["compliance"]))) if metrics["compliance"] else ["puid", "result"])
    _write_rows(wb, "Appendix_LLM_Calls", metrics["llm_calls"], sorted(set().union(*(r.keys() for r in metrics["llm_calls"]))) if metrics["llm_calls"] else ["call_id"])
    _write_rows(wb, "Appendix_LLM_Items", metrics["llm_items"], sorted(set().union(*(r.keys() for r in metrics["llm_items"]))) if metrics["llm_items"] else ["item_id"])
    _write_rows(wb, "Appendix_Prompt_Calls", prompt_calls, ["global_prompt_call_id", "prompt_call_id", "telemetry_stage", "prompt_id", "prompt_name", "prompt_scope", "prompt_category", "section_name", "source_file", "source_function", "model", "provider", "max_output_tokens", "reasoning_effort", "attempt_count", "retry_count", "expected_items", "received_items", "json_valid", "schema_valid", "traceability_ok", "repair_used", "fallback_used", "prompt_success", "prompt_hash", "schema_hash", "payload_hash", "controls_hash", "contract_score", "elapsed_s", "registration_status", "error"])
    # Discovery includes parse errors and unregistered prompt warnings.
    telemetry_discovery = [{"event_type": e.get("event_type", ""), "severity": e.get("severity", "info"), "prompt_id": e.get("prompt_id", ""), "source_file": e.get("source_file", ""), "source_function": e.get("source_function", ""), "message": e.get("message", "")} for e in telemetry_events if e.get("event_type") != "prompt_call"]
    _write_rows(wb, "Appendix_Discovery", discovery_rows + telemetry_discovery, ["event_type", "severity", "prompt_id", "source_file", "source_function", "message"])
    _write_manifest_json(manifest_path, out_dir, summary, package_contents)
    manifest_rows = _manifest_rows(out_dir, raw_path)
    _write_rows(wb, "Appendix_Package_Manifest", manifest_rows, ["artifact", "path", "sha256", "size_bytes"])
    _style_workbook(wb)
    wb.save(output_xlsx)
    # Final PDF and manifest.
    chart_dir = out_dir / "_methodology_charts"
    pdf_rows = {
        "package_contents": package_contents,
        "input_evidence": input_evidence,
        "prompt_contracts": prompt_contracts,
        "prompt_controls": prompt_controls,
        "prompt_run_results": prompt_run_results,
        "llm_validation_summary": llm_validation_summary,
        "prompt_calls": prompt_calls,
    }
    # Refresh metrics with final summary-like values.
    metrics["summary"] = summary
    _generate_pdf(output_pdf, metrics, pdf_rows, chart_dir)
    shutil.rmtree(chart_dir, ignore_errors=True)
    # Recompute manifest rows after PDF exists.
    _write_manifest_json(manifest_path, out_dir, summary, package_contents)
    wb = openpyxl.load_workbook(output_xlsx)
    _write_rows(wb, "Appendix_Package_Manifest", _manifest_rows(out_dir, raw_path), ["artifact", "path", "sha256", "size_bytes"])
    # Update input_evidence with final PDF hash.
    input_evidence = _input_evidence_rows(wb_raw, args, telemetry_components, output_xlsx, output_pdf, manifest_path)
    _write_rows(wb, "input_evidence", input_evidence, ["input_name", "evidence_group", "evidence_role", "artifact_type", "path", "file_count", "sha256", "availability", "purpose", "used_for_stability_comparability"])
    _style_workbook(wb)
    wb.save(output_xlsx)
    print(f"[OK] Run metrics workbook: {output_xlsx}")
    print(f"[OK] Run metrics methodology PDF: {output_pdf}")
    print(f"[OK] Run metrics manifest: {manifest_path}")
    print(f"[OK] Stability-analysis tool: {tool_dst}")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build a user-oriented run-metrics package.")
    parser.add_argument("--raw-metrics", required=True)
    parser.add_argument("--out-dir", required=True)
    parser.add_argument("--prompt-telemetry", action="append", default=[], help="Prompt telemetry JSONL file. Can be supplied more than once.")
    parser.add_argument("--analysis-pack", required=False, default="")
    parser.add_argument("--audit-config", required=False, default="")
    parser.add_argument("--audit-summary-docx", required=False, default="")
    parser.add_argument("--audit-summary-pdf", required=False, default="")
    parser.add_argument("--technical-root", required=False, default="")
    parser.add_argument("--vision360-bundle", required=False, default="")
    parser.add_argument("--trivy-payload", required=False, default="")
    parser.add_argument("--mobsf-report", required=False, default="")
    parser.add_argument("--mobsf-dynamic-report", required=False, default="")
    parser.add_argument("--sast-findings", required=False, default="")
    return parser.parse_args()


if __name__ == "__main__":
    build(parse_args())
